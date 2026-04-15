#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calibrate UK household and dwelling stock totals from official 2024 artifacts.

@author: Max Stoddard
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import asdict, dataclass
from pathlib import Path

from odf.opendocument import load as load_ods
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import load_workbook

from scripts.python.helpers.common.paths import ensure_output_dir


ONS_FAMILIES_HOUSEHOLDS_2024_XLSX_URL = (
    "https://www.ons.gov.uk/file?uri=%2Fpeoplepopulationandcommunity%2Fbirthsdeathsandmarriages%2Ffamilies%2Fdatasets%2F"
    "familiesandhouseholdsfamiliesandhouseholds%2Fcurrent%2Ffamiliesandhouseholdsuk2024.xlsx"
)
ENGLAND_DWELLING_STOCK_2024_ODS_URL = (
    "https://assets.publishing.service.gov.uk/media/682deb00b33f68eaba95391b/LiveTable100.ods"
)
WALES_DWELLING_STOCK_2024_DATASET_URL = "https://stats.gov.wales/en-GB/6476cc20-ddeb-46a5-be64-10a23c8a159f"
SCOTLAND_HOUSEHOLDS_DWELLINGS_2024_XLSX_URL = (
    "https://www.nrscotland.gov.uk/media/nvcaoksr/house-est-24-data.xlsx"
)
NORTHERN_IRELAND_HOUSING_STOCK_2025_XLSX_URL = (
    "https://www.finance-ni.gov.uk/sites/default/files/2025-06/Housing%20Stock%20Tables%202008%20-%202025.xlsx"
)

UK_HOUSEHOLDS_REJECTED_ROUNDED = 28_600_000
ENGLAND_DWELLINGS_REJECTED_ROUNDED = 25_620_000
SCOTLAND_DWELLINGS_REJECTED_ROUNDED = 2_741_000


@dataclass(frozen=True)
class SourceObservation:
    parameter_key: str
    component: str
    published_value: str
    derived_value: int
    units: str
    artifact: str
    source_url: str
    publication_date: str
    extraction_ref: str
    notes: str


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Calibrate UK_HOUSEHOLDS and UK_DWELLINGS from official 2024 ONS, "
            "England, Wales, Scotland, and Northern Ireland source artifacts."
        )
    )
    parser.add_argument(
        "--ons-households-xlsx",
        required=True,
        help="Path to the ONS Families and households in the UK: 2024 XLSX artifact.",
    )
    parser.add_argument(
        "--england-dwellings-ods",
        required=True,
        help="Path to the England Live Table 100 ODS artifact.",
    )
    parser.add_argument(
        "--wales-dwellings-csv",
        required=True,
        help="Path to the Wales dwelling stock CSV download from StatsWales.",
    )
    parser.add_argument(
        "--scotland-dwellings-xlsx",
        required=True,
        help="Path to the Scotland Households and dwellings in Scotland: 2024 XLSX artifact.",
    )
    parser.add_argument(
        "--northern-ireland-dwellings-xlsx",
        required=True,
        help="Path to the Northern Ireland housing stock XLSX artifact.",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Directory where the source-values CSV and calibration-summary JSON will be written.",
    )
    return parser


def _clean_string(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.replace("\xa0", " ").split())
    return str(value)


def _as_int(value: object) -> int:
    if value is None:
        raise ValueError("Expected an integer-like value, found None.")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    cleaned = _clean_string(value).replace(",", "")
    if cleaned in {"", "-", "[x]"}:
        raise ValueError(f"Expected an integer-like value, found {value!r}.")
    return int(float(cleaned))


def _find_header_index(row: list[object], expected_label: str) -> int:
    normalized = _clean_string(expected_label).lower()
    for index, value in enumerate(row):
        if _clean_string(value).lower() == normalized:
            return index
    raise ValueError(f"Could not find header {expected_label!r}.")


def _ods_cell_text(cell: TableCell) -> str:
    parts: list[str] = []
    for paragraph in cell.getElementsByType(P):
        for child in paragraph.childNodes:
            data = getattr(child, "data", None)
            if data:
                parts.append(data)
    return " ".join(" ".join(parts).split())


def _ods_rows(table: Table) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.getElementsByType(TableRow):
        values = [_ods_cell_text(cell) for cell in row.getElementsByType(TableCell)]
        if any(value for value in values):
            rows.append(values)
    return rows


def extract_uk_households(ons_households_xlsx: Path) -> SourceObservation:
    with ons_households_xlsx.open("rb") as handle:
        workbook = load_workbook(handle, data_only=True)
        try:
            worksheet = workbook["5"]
            header = list(next(worksheet.iter_rows(min_row=12, max_row=12, values_only=True)))
            estimate_index = _find_header_index(header, "2024 Estimate")
            for row in worksheet.iter_rows(min_row=13, values_only=True):
                if _clean_string(row[0]) == "All households":
                    estimate_thousands = _as_int(row[estimate_index])
                    return SourceObservation(
                        parameter_key="UK_HOUSEHOLDS",
                        component="United Kingdom households",
                        published_value=f"{estimate_thousands} thousand households",
                        derived_value=estimate_thousands * 1000,
                        units="households",
                        artifact=ons_households_xlsx.name,
                        source_url=ONS_FAMILIES_HOUSEHOLDS_2024_XLSX_URL,
                        publication_date="2025-07-23",
                        extraction_ref="sheet '5' row 'All households' column '2024 Estimate'",
                        notes="The ONS workbook publishes households in thousands, so 28,609 becomes 28,609,000 households.",
                    )
        finally:
            workbook.close()
    raise ValueError("Could not locate the UK all-households row in the ONS workbook.")


def extract_england_dwellings(england_dwellings_ods: Path) -> SourceObservation:
    document = load_ods(str(england_dwellings_ods))
    for table in document.spreadsheet.getElementsByType(Table):
        if table.getAttribute("name") != "2024":
            continue
        rows = _ods_rows(table)
        header = next((row for row in rows if "Area" in row and "Total" in row), None)
        if header is None:
            raise ValueError("Could not find the 2024 England dwelling-stock header row.")
        area_index = header.index("Area")
        total_index = header.index("Total")
        for row in rows:
            if area_index < len(row) and row[area_index] == "England":
                total = _as_int(row[total_index])
                return SourceObservation(
                    parameter_key="UK_DWELLINGS",
                    component="England dwellings",
                    published_value=f"{total}",
                    derived_value=total,
                    units="dwellings",
                    artifact=england_dwellings_ods.name,
                    source_url=ENGLAND_DWELLING_STOCK_2024_ODS_URL,
                    publication_date="2025-05-22",
                    extraction_ref="table '2024' row 'England' column 'Total'",
                    notes=(
                        "The England ODS publishes 25,617,413 in the downloadable table; "
                        "the rounded public headline of 25.62 million is rejected."
                    ),
                )
    raise ValueError("Could not locate the England 2024 table in the ODS artifact.")


def extract_wales_dwellings(wales_dwellings_csv: Path) -> SourceObservation:
    with wales_dwellings_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if (
                row.get("Data description") == "Dwelling stock estimates"
                and row.get("Local authority") == "Wales"
                and row.get("Period") == "31/03/2024"
                and row.get("Tenure") == "All tenures (Number)"
            ):
                total = _as_int(row.get("Data values"))
                return SourceObservation(
                    parameter_key="UK_DWELLINGS",
                    component="Wales dwellings",
                    published_value=f"{total}",
                    derived_value=total,
                    units="dwellings",
                    artifact=wales_dwellings_csv.name,
                    source_url=WALES_DWELLING_STOCK_2024_DATASET_URL,
                    publication_date="2026-01-15",
                    extraction_ref=(
                        "CSV row where Local authority='Wales', Period='31/03/2024', "
                        "Tenure='All tenures (Number)'"
                    ),
                    notes=(
                        "StatsWales metadata states that the estimates are rounded to the nearest 100. "
                        "The published Wales total is 1,482,600."
                    ),
                )
    raise ValueError("Could not locate the Wales 31/03/2024 all-tenures dwelling-stock row.")


def extract_scotland_dwellings(scotland_dwellings_xlsx: Path) -> SourceObservation:
    with scotland_dwellings_xlsx.open("rb") as handle:
        workbook = load_workbook(handle, data_only=True)
        try:
            worksheet = workbook["Table2"]
            header = list(next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True)))
            year_index = _find_header_index(header, "2024")
            for row in worksheet.iter_rows(min_row=5, values_only=True):
                if _clean_string(row[0]) == "Scotland":
                    total = _as_int(row[year_index])
                    return SourceObservation(
                        parameter_key="UK_DWELLINGS",
                        component="Scotland dwellings",
                        published_value=f"{total}",
                        derived_value=total,
                        units="dwellings",
                        artifact=scotland_dwellings_xlsx.name,
                        source_url=SCOTLAND_HOUSEHOLDS_DWELLINGS_2024_XLSX_URL,
                        publication_date="2025-06-26",
                        extraction_ref="sheet 'Table2' row 'Scotland' column '2024'",
                        notes=(
                            "The Scotland workbook notes that figures are rounded to the nearest whole number. "
                            "The published Scotland total is 2,740,973."
                        ),
                    )
        finally:
            workbook.close()
    raise ValueError("Could not locate the Scotland 2024 dwellings row in the workbook.")


def extract_northern_ireland_dwellings(northern_ireland_dwellings_xlsx: Path) -> SourceObservation:
    with northern_ireland_dwellings_xlsx.open("rb") as handle:
        workbook = load_workbook(handle, data_only=True)
        try:
            worksheet = workbook["Table 1.17"]
            header: list[object] | None = None
            for row in worksheet.iter_rows(values_only=True):
                row_values = list(row)
                if "Total Housing Stock" in {_clean_string(value) for value in row_values}:
                    header = row_values
                    break
            if header is None:
                raise ValueError("Could not locate the Table 1.17 header row.")
            total_index = _find_header_index(header, "Total Housing Stock")
            for row in worksheet.iter_rows(values_only=True):
                row_values = list(row)
                if "Northern Ireland" in {_clean_string(value) for value in row_values}:
                    total = _as_int(row_values[total_index])
                    return SourceObservation(
                        parameter_key="UK_DWELLINGS",
                        component="Northern Ireland dwellings",
                        published_value=f"{total}",
                        derived_value=total,
                        units="dwellings",
                        artifact=northern_ireland_dwellings_xlsx.name,
                        source_url=NORTHERN_IRELAND_HOUSING_STOCK_2025_XLSX_URL,
                        publication_date="2025-06-04",
                        extraction_ref="sheet 'Table 1.17' row 'Northern Ireland' column 'Total Housing Stock'",
                        notes="The Northern Ireland workbook publishes the 1 April 2024 stock count directly as 835,988.",
                    )
        finally:
            workbook.close()
    raise ValueError("Could not locate the Northern Ireland 2024 housing-stock row.")


def build_calibration_summary(
    *,
    ons_households_xlsx: Path,
    england_dwellings_ods: Path,
    wales_dwellings_csv: Path,
    scotland_dwellings_xlsx: Path,
    northern_ireland_dwellings_xlsx: Path,
) -> dict[str, object]:
    households = extract_uk_households(ons_households_xlsx)
    england = extract_england_dwellings(england_dwellings_ods)
    wales = extract_wales_dwellings(wales_dwellings_csv)
    scotland = extract_scotland_dwellings(scotland_dwellings_xlsx)
    northern_ireland = extract_northern_ireland_dwellings(northern_ireland_dwellings_xlsx)

    dwelling_components = [england, wales, scotland, northern_ireland]
    selected_uk_dwellings = sum(item.derived_value for item in dwelling_components)
    rejected_mixed_precision = (
        ENGLAND_DWELLINGS_REJECTED_ROUNDED
        + wales.derived_value
        + SCOTLAND_DWELLINGS_REJECTED_ROUNDED
        + northern_ireland.derived_value
    )

    return {
        "methodId": "source_native_downloadable_artifacts_2024",
        "selectedConfigValues": {
            "UK_HOUSEHOLDS": households.derived_value,
            "UK_DWELLINGS": selected_uk_dwellings,
        },
        "sourceObservations": [asdict(households)] + [asdict(item) for item in dwelling_components],
        "selectedAggregation": {
            "UK_HOUSEHOLDS": {
                "component": households.component,
                "publishedValue": households.published_value,
                "derivedValue": households.derived_value,
            },
            "UK_DWELLINGS": {
                "components": [
                    {
                        "component": item.component,
                        "derivedValue": item.derived_value,
                    }
                    for item in dwelling_components
                ],
                "derivedValue": selected_uk_dwellings,
            },
        },
        "rejectedComparisons": [
            {
                "parameterKey": "UK_HOUSEHOLDS",
                "label": "Rounded ONS headline comparator",
                "value": UK_HOUSEHOLDS_REJECTED_ROUNDED,
                "differenceFromSelected": UK_HOUSEHOLDS_REJECTED_ROUNDED - households.derived_value,
                "whyRejected": (
                    "Rejected because the official ONS workbook publishes 28,609 thousand households, "
                    "which converts to 28,609,000 rather than 28,600,000."
                ),
            },
            {
                "parameterKey": "UK_DWELLINGS",
                "label": "Mixed-precision rounded-country comparator",
                "value": rejected_mixed_precision,
                "differenceFromSelected": rejected_mixed_precision - selected_uk_dwellings,
                "formula": (
                    f"{ENGLAND_DWELLINGS_REJECTED_ROUNDED:,} + {wales.derived_value:,} + "
                    f"{SCOTLAND_DWELLINGS_REJECTED_ROUNDED:,} + {northern_ireland.derived_value:,}"
                ),
                "whyRejected": (
                    "Rejected because it mixes rounded England and Scotland public-release headline figures "
                    "with more granular Wales and Northern Ireland downloadable values."
                ),
            },
        ],
        "artifacts": {
            "onsHouseholdsXlsx": ons_households_xlsx.name,
            "englandDwellingsOds": england_dwellings_ods.name,
            "walesDwellingsCsv": wales_dwellings_csv.name,
            "scotlandDwellingsXlsx": scotland_dwellings_xlsx.name,
            "northernIrelandDwellingsXlsx": northern_ireland_dwellings_xlsx.name,
        },
    }


def _write_source_values_csv(output_path: Path, summary: dict[str, object]) -> None:
    selected_config_values = summary["selectedConfigValues"]
    rejected_comparisons = summary["rejectedComparisons"]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "row_type",
                "parameter_key",
                "component",
                "published_value",
                "derived_value",
                "units",
                "artifact",
                "source_url",
                "publication_date",
                "extraction_ref",
                "notes",
            ]
        )
        for observation in summary["sourceObservations"]:
            writer.writerow(
                [
                    "source_observation",
                    observation["parameter_key"],
                    observation["component"],
                    observation["published_value"],
                    observation["derived_value"],
                    observation["units"],
                    observation["artifact"],
                    observation["source_url"],
                    observation["publication_date"],
                    observation["extraction_ref"],
                    observation["notes"],
                ]
            )
        writer.writerow(
            [
                "selected_total",
                "UK_HOUSEHOLDS",
                "Selected config value",
                "",
                selected_config_values["UK_HOUSEHOLDS"],
                "households",
                "",
                "",
                "",
                "",
                "Selected source-native ONS workbook value.",
            ]
        )
        writer.writerow(
            [
                "selected_total",
                "UK_DWELLINGS",
                "Selected config value",
                "",
                selected_config_values["UK_DWELLINGS"],
                "dwellings",
                "",
                "",
                "",
                "",
                "Selected source-native sum of England, Wales, Scotland, and Northern Ireland downloadable totals.",
            ]
        )
        for rejected in rejected_comparisons:
            writer.writerow(
                [
                    "rejected_comparison",
                    rejected["parameterKey"],
                    rejected["label"],
                    "",
                    rejected["value"],
                    "households" if rejected["parameterKey"] == "UK_HOUSEHOLDS" else "dwellings",
                    "",
                    "",
                    "",
                    rejected.get("formula", ""),
                    rejected["whyRejected"],
                ]
            )


def _write_summary_json(output_path: Path, summary: dict[str, object]) -> None:
    output_path.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")


def run_calibration(
    *,
    ons_households_xlsx: Path,
    england_dwellings_ods: Path,
    wales_dwellings_csv: Path,
    scotland_dwellings_xlsx: Path,
    northern_ireland_dwellings_xlsx: Path,
    output_dir: Path,
) -> dict[str, object]:
    summary = build_calibration_summary(
        ons_households_xlsx=ons_households_xlsx,
        england_dwellings_ods=england_dwellings_ods,
        wales_dwellings_csv=wales_dwellings_csv,
        scotland_dwellings_xlsx=scotland_dwellings_xlsx,
        northern_ireland_dwellings_xlsx=northern_ireland_dwellings_xlsx,
    )
    output_root = ensure_output_dir(output_dir)
    _write_source_values_csv(output_root / "UkHousingStockTotals2024SourceValues.csv", summary)
    _write_summary_json(output_root / "UkHousingStockTotals2024CalibrationSummary.json", summary)
    return summary


def main() -> None:
    args = build_arg_parser().parse_args()
    summary = run_calibration(
        ons_households_xlsx=Path(args.ons_households_xlsx),
        england_dwellings_ods=Path(args.england_dwellings_ods),
        wales_dwellings_csv=Path(args.wales_dwellings_csv),
        scotland_dwellings_xlsx=Path(args.scotland_dwellings_xlsx),
        northern_ireland_dwellings_xlsx=Path(args.northern_ireland_dwellings_xlsx),
        output_dir=Path(args.output_dir),
    )
    selected = summary["selectedConfigValues"]
    print(f"UK_HOUSEHOLDS = {selected['UK_HOUSEHOLDS']}")
    print(f"UK_DWELLINGS = {selected['UK_DWELLINGS']}")


if __name__ == "__main__":
    main()
