#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calibrate RENT_GROSS_YIELD from official 2024 rent and house-price sources.

@author: Max Stoddard
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import fmean

from openpyxl import load_workbook

from scripts.python.helpers.common.cli import format_float
from scripts.python.helpers.common.paths import ensure_output_dir, repo_root


RENT_GROSS_YIELD_KEY = "RENT_GROSS_YIELD"
PIPR_SOURCE_URL = (
    "https://www.ons.gov.uk/file?uri=/economy/inflationandpriceindices/datasets/"
    "priceindexofprivaterentsukmonthlypricestatistics/18december2024/"
    "priceindexofprivaterentsukmonthlypricestatistics.xlsx"
)
HPI_FULL_FILE_SOURCE_URL = (
    "https://publicdata.landregistry.gov.uk/market-trend-data/house-price-index-data/"
    "UK-HPI-full-file-2024-12.csv"
)
AVERAGE_PRICES_SOURCE_URL = (
    "https://publicdata.landregistry.gov.uk/market-trend-data/house-price-index-data/"
    "Average-prices-2024-12.csv"
)
ONS_PIPR_RELEASE_URL = "https://www.ons.gov.uk/releases/privaterentandhousepricesukdecember2024"
ONS_PIPR_METHODOLOGY_URL = (
    "https://www.ons.gov.uk/economy/inflationandpriceindices/methodologies/"
    "priceindexofprivaterentsdetailedmethodology"
)
GOVUK_HPI_DOWNLOADS_URL = (
    "https://www.gov.uk/government/statistical-data-sets/"
    "uk-house-price-index-data-downloads-december-2024"
)

PIPR_TABLE_NAME = "Table 1"
PIPR_HEADER_ROW = 3
PIPR_RENT_GEOGRAPHY = "Great Britain"
PIPR_UNAVAILABLE_RENT_GEOGRAPHIES = ("UK", "Northern Ireland")
HPI_REGION_NAME_UK = "United Kingdom"
TARGET_YEAR = 2024
MONTHS_IN_YEAR = 12
EXPECTED_PIPR_RENT_MONTHS = tuple(range(1, 12))
SUMMARY_FILE_NAME = "RentGrossYield2024Summary.json"
SOURCE_VALUES_FILE_NAME = "RentGrossYield2024SourceValues.csv"


@dataclass(frozen=True)
class MonthlyValue:
    source: str
    geography: str
    date: str
    value: float
    units: str
    artifact: str
    source_url: str
    extraction_ref: str
    notes: str


@dataclass(frozen=True)
class UnavailableRentLevel:
    geography: str
    months_observed: list[str]
    raw_values: list[str]
    artifact: str
    source_url: str
    extraction_ref: str
    notes: str


@dataclass(frozen=True)
class SourceArtifact:
    label: str
    path: str
    sha256: str
    source_url: str
    notes: str


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Calibrate RENT_GROSS_YIELD as 12 * average monthly rent / average house price "
            "from the December 2024 ONS PIPR and HM Land Registry UK HPI artifacts."
        )
    )
    parser.add_argument("--pipr-xlsx", required=True, help="Path to the December 2024 ONS PIPR XLSX artifact.")
    parser.add_argument(
        "--hpi-full-csv",
        required=True,
        help="Path to the December 2024 HM Land Registry UK HPI full-file CSV artifact.",
    )
    parser.add_argument(
        "--average-prices-csv",
        default=None,
        help="Optional path to the December 2024 HM Land Registry Average-prices CSV comparator.",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help=f"Directory where {SOURCE_VALUES_FILE_NAME} and {SUMMARY_FILE_NAME} will be written.",
    )
    parser.add_argument(
        "--target-year",
        type=int,
        default=TARGET_YEAR,
        help=f"Calendar year to extract. Defaults to {TARGET_YEAR}.",
    )
    return parser


def _clean_string(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _as_float(value: object) -> float:
    if _is_number(value):
        return float(value)
    cleaned = _clean_string(value).replace(",", "")
    if not cleaned:
        raise ValueError("Expected a numeric value, found blank text.")
    return float(cleaned)


def _cell_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_string(value)
    for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Could not parse date value {value!r}.")


def _find_header_index(header: list[object], expected_label: str) -> int:
    normalized = expected_label.lower()
    for index, value in enumerate(header):
        if _clean_string(value).lower() == normalized:
            return index
    raise ValueError(f"Could not find header {expected_label!r}.")


def _sha256(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _repo_relative(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(repo_root()))
    except ValueError:
        return str(resolved)


def _artifact_record(path: Path, *, label: str, source_url: str, notes: str) -> SourceArtifact:
    return SourceArtifact(
        label=label,
        path=_repo_relative(path),
        sha256=_sha256(path),
        source_url=source_url,
        notes=notes,
    )


def _load_pipr_rows(pipr_xlsx: Path) -> tuple[list[object], list[tuple[object, ...]]]:
    with pipr_xlsx.open("rb") as handle:
        workbook = load_workbook(handle, read_only=True, data_only=True)
        try:
            worksheet = workbook[PIPR_TABLE_NAME]
            header = list(next(worksheet.iter_rows(min_row=PIPR_HEADER_ROW, max_row=PIPR_HEADER_ROW, values_only=True)))
            rows = [tuple(row) for row in worksheet.iter_rows(min_row=PIPR_HEADER_ROW + 1, values_only=True)]
        finally:
            workbook.close()
    return header, rows


def extract_pipr_rent_values(
    pipr_xlsx: Path,
    *,
    target_year: int = TARGET_YEAR,
    geography: str = PIPR_RENT_GEOGRAPHY,
) -> list[MonthlyValue]:
    header, rows = _load_pipr_rows(pipr_xlsx)
    time_index = _find_header_index(header, "Time period")
    area_index = _find_header_index(header, "Area name")
    rent_index = _find_header_index(header, "Rental price")

    observations: list[MonthlyValue] = []
    for row in rows:
        if _clean_string(row[area_index]) != geography:
            continue
        observed_date = _cell_date(row[time_index])
        if observed_date.year != target_year:
            continue
        raw_value = row[rent_index]
        if not _is_number(raw_value):
            continue
        observations.append(
            MonthlyValue(
                source="ONS PIPR monthly price statistics",
                geography=geography,
                date=observed_date.isoformat(),
                value=float(raw_value),
                units="GBP per month",
                artifact=pipr_xlsx.name,
                source_url=PIPR_SOURCE_URL,
                extraction_ref=(
                    f"sheet {PIPR_TABLE_NAME!r}, Area name={geography!r}, "
                    f"year={target_year}, column 'Rental price'"
                ),
                notes=(
                    "Selected Great Britain because the requested December 2024 PIPR workbook "
                    "does not publish UK rent-price levels."
                ),
            )
        )

    observations.sort(key=lambda item: item.date)
    if not observations:
        raise ValueError(f"No numeric PIPR rent-price values found for {geography!r} in {target_year}.")
    observed_months = [datetime.fromisoformat(item.date).month for item in observations]
    if observed_months != list(EXPECTED_PIPR_RENT_MONTHS):
        raise ValueError(
            f"Expected PIPR rent-price values for months {list(EXPECTED_PIPR_RENT_MONTHS)} "
            f"in {target_year}, found months {observed_months}."
        )
    return observations


def extract_unavailable_pipr_rent_levels(
    pipr_xlsx: Path,
    *,
    target_year: int = TARGET_YEAR,
    geographies: tuple[str, ...] = PIPR_UNAVAILABLE_RENT_GEOGRAPHIES,
) -> list[UnavailableRentLevel]:
    header, rows = _load_pipr_rows(pipr_xlsx)
    time_index = _find_header_index(header, "Time period")
    area_index = _find_header_index(header, "Area name")
    rent_index = _find_header_index(header, "Rental price")

    grouped: dict[str, dict[str, set[str] | list[str]]] = {
        geography: {"months": [], "raw_values": set()} for geography in geographies
    }
    for row in rows:
        geography = _clean_string(row[area_index])
        if geography not in grouped:
            continue
        observed_date = _cell_date(row[time_index])
        if observed_date.year != target_year:
            continue
        raw_value = row[rent_index]
        if _is_number(raw_value):
            continue
        grouped[geography]["months"].append(observed_date.isoformat())  # type: ignore[union-attr]
        grouped[geography]["raw_values"].add(_clean_string(raw_value))  # type: ignore[union-attr]

    results: list[UnavailableRentLevel] = []
    for geography, values in grouped.items():
        months = sorted(values["months"])  # type: ignore[arg-type]
        if not months:
            continue
        results.append(
            UnavailableRentLevel(
                geography=geography,
                months_observed=months,
                raw_values=sorted(values["raw_values"]),  # type: ignore[arg-type]
                artifact=pipr_xlsx.name,
                source_url=PIPR_SOURCE_URL,
                extraction_ref=(
                    f"sheet {PIPR_TABLE_NAME!r}, Area name={geography!r}, "
                    f"year={target_year}, column 'Rental price'"
                ),
                notes=(
                    "The requested December 2024 PIPR workbook marks this rent-price level as unavailable, "
                    "so it is not used in the selected numerator."
                ),
            )
        )
    return results


def extract_hpi_uk_average_prices(
    hpi_full_csv: Path,
    *,
    target_year: int = TARGET_YEAR,
) -> list[MonthlyValue]:
    with hpi_full_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"Missing header row in {hpi_full_csv}.")
        field_map = {field.strip(): field for field in reader.fieldnames if field}
        date_field = field_map.get("Date")
        region_field = field_map.get("RegionName")
        price_field = field_map.get("AveragePrice")
        if date_field is None or region_field is None or price_field is None:
            raise ValueError(f"Missing required HPI full-file columns in {hpi_full_csv}.")

        observations: list[MonthlyValue] = []
        for row in reader:
            if _clean_string(row[region_field]) != HPI_REGION_NAME_UK:
                continue
            observed_date = _cell_date(row[date_field])
            if observed_date.year != target_year:
                continue
            raw_price = row.get(price_field, "")
            if not _clean_string(raw_price):
                continue
            observations.append(
                MonthlyValue(
                    source="HM Land Registry UK HPI full file",
                    geography=HPI_REGION_NAME_UK,
                    date=observed_date.isoformat(),
                    value=_as_float(raw_price),
                    units="GBP",
                    artifact=hpi_full_csv.name,
                    source_url=HPI_FULL_FILE_SOURCE_URL,
                    extraction_ref=(
                        f"RegionName={HPI_REGION_NAME_UK!r}, year={target_year}, "
                        "column 'AveragePrice'"
                    ),
                    notes="Selected unadjusted United Kingdom monthly AveragePrice from the richer HPI full file.",
                )
            )

    observations.sort(key=lambda item: item.date)
    if len(observations) != MONTHS_IN_YEAR:
        raise ValueError(
            f"Expected {MONTHS_IN_YEAR} UK HPI AveragePrice rows for {target_year}, "
            f"found {len(observations)}."
        )
    return observations


def inspect_average_prices_comparator(
    average_prices_csv: Path,
    *,
    target_year: int = TARGET_YEAR,
) -> dict[str, object]:
    with average_prices_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"Missing header row in {average_prices_csv}.")
        fieldnames = [field for field in reader.fieldnames if field]
        field_set = set(fieldnames)
        has_weight_fields = any(
            field in field_set for field in ("SalesVolume", "CashSalesVolume", "MortgageSalesVolume")
        )
        matching_rows = 0
        for row in reader:
            date_text = row.get("Date", "")
            if not date_text:
                continue
            try:
                observed_date = _cell_date(date_text)
            except ValueError:
                continue
            if observed_date.year != target_year:
                continue
            region = _clean_string(row.get("Region_Name"))
            area_code = _clean_string(row.get("Area_Code"))
            if region == HPI_REGION_NAME_UK or area_code == "K02000001":
                matching_rows += 1

    return {
        "artifact": average_prices_csv.name,
        "sourceUrl": AVERAGE_PRICES_SOURCE_URL,
        "fieldNames": fieldnames,
        "targetYearUkRows": matching_rows,
        "hasWeightFields": has_weight_fields,
        "status": "rejected_comparator",
        "whyRejected": (
            "Rejected for the selected calibration because the HPI full file contains the same selected "
            "AveragePrice field plus richer audit columns and is already the repo's validation-source shape. "
            "No extra monthly SalesVolume weighting is applied because RENT_GROSS_YIELD needs a price-level "
            "denominator rather than a transaction-flow-weighted denominator."
        ),
    }


def build_calibration_summary(
    *,
    pipr_xlsx: Path,
    hpi_full_csv: Path,
    average_prices_csv: Path | None = None,
    target_year: int = TARGET_YEAR,
) -> dict[str, object]:
    rent_values = extract_pipr_rent_values(pipr_xlsx, target_year=target_year)
    unavailable_rent_levels = extract_unavailable_pipr_rent_levels(pipr_xlsx, target_year=target_year)
    hpi_values = extract_hpi_uk_average_prices(hpi_full_csv, target_year=target_year)

    average_monthly_rent = fmean(item.value for item in rent_values)
    average_house_price = fmean(item.value for item in hpi_values)
    selected_value = round(MONTHS_IN_YEAR * average_monthly_rent / average_house_price, 10)

    artifacts = [
        asdict(
            _artifact_record(
                pipr_xlsx,
                label="ONS PIPR monthly price statistics, December 2024 edition",
                source_url=PIPR_SOURCE_URL,
                notes="Source workbook for Great Britain monthly rent-price levels.",
            )
        ),
        asdict(
            _artifact_record(
                hpi_full_csv,
                label="HM Land Registry UK HPI full file, December 2024 edition",
                source_url=HPI_FULL_FILE_SOURCE_URL,
                notes="Selected source for the United Kingdom monthly AveragePrice denominator.",
            )
        ),
    ]
    average_prices_comparator = None
    if average_prices_csv is not None:
        artifacts.append(
            asdict(
                _artifact_record(
                    average_prices_csv,
                    label="HM Land Registry Average-prices file, December 2024 edition",
                    source_url=AVERAGE_PRICES_SOURCE_URL,
                    notes="Retained comparator requested for source-choice audit.",
                )
            )
        )
        average_prices_comparator = inspect_average_prices_comparator(
            average_prices_csv,
            target_year=target_year,
        )

    return {
        "methodId": "pipr_gb_mean_hmlr_uk_full_file_mean_2024",
        "methodRationale": (
            "RENT_GROSS_YIELD initializes expected gross BTL rental yield in the model. The selected calculation "
            "uses the requested PIPR/HPI rent-to-price formula while avoiding an unsupported UK rent-price level: "
            "the December 2024 PIPR workbook publishes Great Britain rent-price levels through November 2024, "
            "but marks UK and Northern Ireland rent-price levels as unavailable. The HPI denominator uses the "
            "United Kingdom unadjusted AveragePrice series from the full file for all twelve 2024 months."
        ),
        "selectedConfigValues": {RENT_GROSS_YIELD_KEY: selected_value},
        "calculation": {
            "targetYear": target_year,
            "formula": "12 * mean(PIPR Great Britain monthly rent) / mean(HMLR UK HPI AveragePrice)",
            "rentMean": average_monthly_rent,
            "rentMeanMonths": [item.date for item in rent_values],
            "rentMeanGeography": PIPR_RENT_GEOGRAPHY,
            "housePriceMean": average_house_price,
            "housePriceMeanMonths": [item.date for item in hpi_values],
            "housePriceGeography": HPI_REGION_NAME_UK,
            "unroundedValue": MONTHS_IN_YEAR * average_monthly_rent / average_house_price,
            "selectedConfigValue": selected_value,
        },
        "sourceValues": {
            "piprRentValues": [asdict(item) for item in rent_values],
            "hpiAveragePrices": [asdict(item) for item in hpi_values],
            "unavailablePiprRentLevels": [asdict(item) for item in unavailable_rent_levels],
        },
        "sourceArtifacts": artifacts,
        "sourceReferences": {
            "onsPiprRelease": ONS_PIPR_RELEASE_URL,
            "onsPiprMethodology": ONS_PIPR_METHODOLOGY_URL,
            "govukHpiDownloads": GOVUK_HPI_DOWNLOADS_URL,
        },
        "rejectedComparisons": [
            item
            for item in [
                {
                    "label": "Literal UK PIPR rent-price numerator",
                    "status": "rejected_unavailable",
                    "whyRejected": (
                        "The requested December 2024 PIPR workbook provides UK index rows, but the 'Rental price' "
                        "column is marked unavailable for UK rows."
                    ),
                },
                {
                    "label": "Transaction-volume-weighted monthly HPI denominator",
                    "status": "rejected_method_mismatch",
                    "whyRejected": (
                        "The December 2024 full file lacks SalesVolume for November and December UK rows, and "
                        "transaction-volume weighting would convert a price-level denominator into a flow-weighted "
                        "denominator. RENT_GROSS_YIELD is better represented by the full-year price-level mean."
                    ),
                },
                average_prices_comparator,
            ]
            if item is not None
        ],
    }


def _write_summary_json(output_path: Path, summary: dict[str, object]) -> None:
    output_path.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")


def _write_source_values_csv(output_path: Path, summary: dict[str, object]) -> None:
    source_values = summary["sourceValues"]
    calculation = summary["calculation"]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "row_type",
                "source",
                "geography",
                "date",
                "value",
                "units",
                "artifact",
                "source_url",
                "extraction_ref",
                "notes",
            ]
        )
        for row_type, key in (
            ("rent_observation", "piprRentValues"),
            ("house_price_observation", "hpiAveragePrices"),
        ):
            for item in source_values[key]:
                writer.writerow(
                    [
                        row_type,
                        item["source"],
                        item["geography"],
                        item["date"],
                        item["value"],
                        item["units"],
                        item["artifact"],
                        item["source_url"],
                        item["extraction_ref"],
                        item["notes"],
                    ]
                )
        for item in source_values["unavailablePiprRentLevels"]:
            writer.writerow(
                [
                    "unavailable_rent_level",
                    "ONS PIPR monthly price statistics",
                    item["geography"],
                    ";".join(item["months_observed"]),
                    ";".join(item["raw_values"]),
                    "GBP per month",
                    item["artifact"],
                    item["source_url"],
                    item["extraction_ref"],
                    item["notes"],
                ]
            )
        writer.writerow(
            [
                "selected_value",
                "derived",
                f"{calculation['rentMeanGeography']} rent / {calculation['housePriceGeography']} price",
                calculation["targetYear"],
                calculation["selectedConfigValue"],
                "fraction",
                "",
                "",
                calculation["formula"],
                (
                    f"rentMean={calculation['rentMean']}; "
                    f"housePriceMean={calculation['housePriceMean']}; "
                    f"unroundedValue={calculation['unroundedValue']}"
                ),
            ]
        )


def run_calibration(
    *,
    pipr_xlsx: Path,
    hpi_full_csv: Path,
    average_prices_csv: Path | None,
    output_dir: Path,
    target_year: int = TARGET_YEAR,
) -> dict[str, object]:
    summary = build_calibration_summary(
        pipr_xlsx=pipr_xlsx,
        hpi_full_csv=hpi_full_csv,
        average_prices_csv=average_prices_csv,
        target_year=target_year,
    )
    output_root = ensure_output_dir(output_dir)
    _write_source_values_csv(output_root / SOURCE_VALUES_FILE_NAME, summary)
    _write_summary_json(output_root / SUMMARY_FILE_NAME, summary)
    return summary


def main() -> None:
    args = build_arg_parser().parse_args()
    summary = run_calibration(
        pipr_xlsx=Path(args.pipr_xlsx),
        hpi_full_csv=Path(args.hpi_full_csv),
        average_prices_csv=Path(args.average_prices_csv) if args.average_prices_csv else None,
        output_dir=Path(args.output_dir),
        target_year=args.target_year,
    )
    selected_value = summary["selectedConfigValues"][RENT_GROSS_YIELD_KEY]
    print(f"{RENT_GROSS_YIELD_KEY} = {format_float(selected_value)}")


if __name__ == "__main__":
    main()
