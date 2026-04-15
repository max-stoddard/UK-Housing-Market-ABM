from __future__ import annotations

import csv
import json
import tempfile
import unittest
import warnings
from pathlib import Path

from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import Workbook

from scripts.python.calibration.official.uk_housing_stock_totals_2024 import (
    build_calibration_summary,
    extract_england_dwellings,
    extract_northern_ireland_dwellings,
    extract_scotland_dwellings,
    extract_uk_households,
    extract_wales_dwellings,
    run_calibration,
)

warnings.simplefilter("ignore", ResourceWarning)


def _add_ods_row(table: Table, values: list[object]) -> None:
    row = TableRow()
    for value in values:
        cell = TableCell()
        if value is not None:
            cell.addElement(P(text=str(value)))
        row.addElement(cell)
    table.addElement(row)


class TestUkHousingStockTotals2024(unittest.TestCase):
    def test_extract_uk_households_reads_thousands_and_converts_to_households(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "ons.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "5"
            worksheet.cell(row=12, column=1, value="Number of households")
            worksheet.cell(row=12, column=2, value="2024 Estimate")
            worksheet.cell(row=13, column=1, value="All households")
            worksheet.cell(row=13, column=2, value=28609)
            workbook.save(workbook_path)
            workbook.close()

            observation = extract_uk_households(workbook_path)

        self.assertEqual(observation.derived_value, 28_609_000)
        self.assertEqual(observation.published_value, "28609 thousand households")

    def test_extract_england_dwellings_reads_total_from_ods(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            ods_path = Path(tmp_dir) / "england.ods"
            document = OpenDocumentSpreadsheet()
            table = Table(name="2024")
            _add_ods_row(table, ["ignored"])
            _add_ods_row(
                table,
                ["Old ONS code", "New ONS code", "Area", "Local authority", "Private sector", "Total"],
            )
            _add_ods_row(table, ["", "E92000001", "England", "1,574,098", "21,342,624", "25,617,413"])
            document.spreadsheet.addElement(table)
            document.save(str(ods_path))

            observation = extract_england_dwellings(ods_path)

        self.assertEqual(observation.derived_value, 25_617_413)

    def test_extract_wales_dwellings_reads_statswales_csv_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            csv_path = Path(tmp_dir) / "wales.csv"
            with csv_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "Data values",
                        "Data description",
                        "Local authority",
                        "Period",
                        "Tenure",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "Data values": "1482600",
                        "Data description": "Dwelling stock estimates",
                        "Local authority": "Wales",
                        "Period": "31/03/2024",
                        "Tenure": "All tenures (Number)",
                    }
                )

            observation = extract_wales_dwellings(csv_path)

        self.assertEqual(observation.derived_value, 1_482_600)

    def test_extract_scotland_dwellings_reads_table2_2024_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "scotland.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Table2"
            worksheet.cell(row=4, column=1, value="Area Name")
            worksheet.cell(row=4, column=2, value="2024")
            worksheet.cell(row=5, column=1, value="Scotland")
            worksheet.cell(row=5, column=2, value=2740973)
            workbook.save(workbook_path)
            workbook.close()

            observation = extract_scotland_dwellings(workbook_path)

        self.assertEqual(observation.derived_value, 2_740_973)

    def test_extract_northern_ireland_dwellings_reads_table_1_17_total(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "ni.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Table 1.17"
            worksheet.cell(row=10, column=1, value="Area")
            worksheet.cell(row=10, column=2, value="Total Housing Stock")
            worksheet.cell(row=11, column=1, value="Northern Ireland")
            worksheet.cell(row=11, column=2, value=835988)
            workbook.save(workbook_path)
            workbook.close()

            observation = extract_northern_ireland_dwellings(workbook_path)

        self.assertEqual(observation.derived_value, 835_988)

    def test_build_calibration_summary_selects_source_native_values_and_rejects_mixed_precision(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            paths = self._build_synthetic_artifacts(Path(tmp_dir))

            summary = build_calibration_summary(**paths)

        self.assertEqual(summary["selectedConfigValues"]["UK_HOUSEHOLDS"], 28_609_000)
        self.assertEqual(summary["selectedConfigValues"]["UK_DWELLINGS"], 30_676_974)
        rejected_dwellings = next(
            item
            for item in summary["rejectedComparisons"]
            if item["parameterKey"] == "UK_DWELLINGS"
        )
        self.assertEqual(rejected_dwellings["value"], 30_679_588)
        self.assertEqual(rejected_dwellings["differenceFromSelected"], 2_614)

    def test_run_calibration_writes_source_values_csv_and_summary_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_root = Path(tmp_dir)
            paths = self._build_synthetic_artifacts(temp_root)
            output_dir = temp_root / "output"

            summary = run_calibration(output_dir=output_dir, **paths)

            source_values_path = output_dir / "UkHousingStockTotals2024SourceValues.csv"
            summary_path = output_dir / "UkHousingStockTotals2024CalibrationSummary.json"
            self.assertTrue(source_values_path.exists())
            self.assertTrue(summary_path.exists())
            loaded_summary = json.loads(summary_path.read_text(encoding="utf-8"))
            self.assertEqual(loaded_summary["selectedConfigValues"], summary["selectedConfigValues"])

    @staticmethod
    def _build_synthetic_artifacts(temp_root: Path) -> dict[str, Path]:
        ons_path = temp_root / "ons.xlsx"
        england_path = temp_root / "england.ods"
        wales_path = temp_root / "wales.csv"
        scotland_path = temp_root / "scotland.xlsx"
        ni_path = temp_root / "ni.xlsx"

        ons_workbook = Workbook()
        ons_sheet = ons_workbook.active
        ons_sheet.title = "5"
        ons_sheet.cell(row=12, column=1, value="Number of households")
        ons_sheet.cell(row=12, column=2, value="2024 Estimate")
        ons_sheet.cell(row=13, column=1, value="All households")
        ons_sheet.cell(row=13, column=2, value=28609)
        ons_workbook.save(ons_path)
        ons_workbook.close()

        england_document = OpenDocumentSpreadsheet()
        england_table = Table(name="2024")
        _add_ods_row(england_table, ["ignored"])
        _add_ods_row(
            england_table,
            ["Old ONS code", "New ONS code", "Area", "Local authority", "Private sector", "Total"],
        )
        _add_ods_row(england_table, ["", "E92000001", "England", "1,574,098", "21,342,624", "25,617,413"])
        england_document.spreadsheet.addElement(england_table)
        england_document.save(str(england_path))

        with wales_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Data values",
                    "Data description",
                    "Local authority",
                    "Period",
                    "Tenure",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "Data values": "1482600",
                    "Data description": "Dwelling stock estimates",
                    "Local authority": "Wales",
                    "Period": "31/03/2024",
                    "Tenure": "All tenures (Number)",
                }
            )

        scotland_workbook = Workbook()
        scotland_sheet = scotland_workbook.active
        scotland_sheet.title = "Table2"
        scotland_sheet.cell(row=4, column=1, value="Area Name")
        scotland_sheet.cell(row=4, column=2, value="2024")
        scotland_sheet.cell(row=5, column=1, value="Scotland")
        scotland_sheet.cell(row=5, column=2, value=2740973)
        scotland_workbook.save(scotland_path)
        scotland_workbook.close()

        ni_workbook = Workbook()
        ni_sheet = ni_workbook.active
        ni_sheet.title = "Table 1.17"
        ni_sheet.cell(row=10, column=1, value="Area")
        ni_sheet.cell(row=10, column=2, value="Total Housing Stock")
        ni_sheet.cell(row=11, column=1, value="Northern Ireland")
        ni_sheet.cell(row=11, column=2, value=835988)
        ni_workbook.save(ni_path)
        ni_workbook.close()

        return {
            "ons_households_xlsx": ons_path,
            "england_dwellings_ods": england_path,
            "wales_dwellings_csv": wales_path,
            "scotland_dwellings_xlsx": scotland_path,
            "northern_ireland_dwellings_xlsx": ni_path,
        }


if __name__ == "__main__":
    unittest.main()
