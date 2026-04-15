from __future__ import annotations

import csv
import subprocess
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from scripts.python.calibration.boe.boe_bank_parameter_calibration import build_arg_parser as build_calibration_arg_parser
from scripts.python.helpers.boe.bank_parameters import (
    build_bank_rate_monthly_series,
    build_method_search_output,
    extract_housing_tools_spread_series,
    load_bank_rate_history,
    load_vtuz_series,
    per_household_credit,
)
from scripts.python.experiments.boe.boe_bank_parameter_method_search import build_arg_parser as build_experiment_arg_parser


class TestBoeBankParameterCalibration(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.repo_root = Path(__file__).resolve().parents[3]
        cls.calibration_module = "scripts.python.calibration.boe.boe_bank_parameter_calibration"

    def test_experiment_parser_defaults(self) -> None:
        args = build_experiment_arg_parser().parse_args(
            [
                "--bank-rate-csv",
                "bank-rate.csv",
                "--housing-tools-xlsx",
                "housing-tools.xlsx",
                "--vtuz-csv",
                "vtuz.csv",
                "--ons-households",
                "28600000",
            ]
        )
        self.assertEqual(args.target_year, 2024)
        self.assertIsNone(args.output_dir)

    def test_calibration_parser_requires_output_dir(self) -> None:
        args = build_calibration_arg_parser().parse_args(
            [
                "--bank-rate-csv",
                "bank-rate.csv",
                "--housing-tools-xlsx",
                "housing-tools.xlsx",
                "--vtuz-csv",
                "vtuz.csv",
                "--ons-households",
                "28600000",
                "--output-dir",
                "evidence",
            ]
        )
        self.assertEqual(args.target_year, 2024)
        self.assertEqual(args.output_dir, "evidence")

    def test_bank_rate_history_to_monthly_series_is_daily_weighted(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            csv_path = Path(tmp_dir) / "bank-rate.csv"
            csv_path.write_text(
                "Date Changed,Rate\n"
                "01 Dec 23,2.00\n"
                "01 Jan 24,3.00\n"
                "15 Jan 24,4.00\n",
                encoding="utf-8",
            )
            events = load_bank_rate_history(csv_path)
            monthly = build_bank_rate_monthly_series(events, year=2024)

        january = monthly[0]
        february = monthly[1]
        expected_january = ((14 * 0.03) + (17 * 0.04)) / 31.0
        self.assertAlmostEqual(january.value, expected_january, places=12)
        self.assertAlmostEqual(february.value, 0.04, places=12)

    def test_extract_housing_tools_spread_series_reads_target_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "housing-tools.xlsx"
            workbook = Workbook()
            default_sheet = workbook.active
            default_sheet.title = "Ignore"
            worksheet = workbook.create_sheet("8. Spreads new mortgage lending")
            for _ in range(4):
                worksheet.append([None, None])
            worksheet.append([date(2024, 1, 31), 0.5])
            worksheet.append([date(2024, 2, 29), 0.6])
            workbook.save(workbook_path)

            observations = extract_housing_tools_spread_series(workbook_path)

        self.assertEqual(len(observations), 2)
        self.assertEqual(observations[0].observation_date, date(2024, 1, 31))
        self.assertAlmostEqual(observations[1].value, 0.6, places=12)

    def test_vtuz_series_converts_to_per_household_credit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            csv_path = Path(tmp_dir) / "vtuz.csv"
            csv_path.write_text(
                "DATE,LPMVTUZ\n"
                "31 Jan 2024,286\n"
                "29 Feb 2024,572\n",
                encoding="utf-8",
            )
            vtuz = load_vtuz_series(csv_path)
            credit = per_household_credit(vtuz, households=1_000_000)

        self.assertEqual(len(credit), 2)
        self.assertAlmostEqual(credit[0].value, 286.0, places=12)
        self.assertAlmostEqual(credit[1].value, 572.0, places=12)

    def test_method_search_rejects_negative_2024_fit_but_selects_positive_longer_fit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            bank_rate_csv = tmp_path / "bank-rate.csv"
            housing_tools_xlsx = tmp_path / "housing-tools.xlsx"
            vtuz_csv = tmp_path / "vtuz.csv"

            bank_rate_csv.write_text(
                "Date Changed,Rate\n"
                "01 Dec 22,2.00\n"
                "01 Jan 24,3.00\n",
                encoding="utf-8",
            )

            workbook = Workbook()
            workbook.remove(workbook.active)
            worksheet = workbook.create_sheet("8. Spreads new mortgage lending")
            for _ in range(4):
                worksheet.append([None, None])
            spread_rows = [
                (date(2023, 10, 31), 1.0),
                (date(2023, 11, 30), 2.0),
                (date(2023, 12, 31), 3.0),
            ]
            spread_rows.extend(
                (
                    self._month_end_2024(month),
                    2.9 - (0.1 * (month - 1)),
                )
                for month in range(1, 13)
            )
            for observation_date, value in spread_rows:
                worksheet.append([observation_date, value])
            workbook.save(housing_tools_xlsx)

            with vtuz_csv.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["DATE", "LPMVTUZ"])
                writer.writerow(["31 Oct 2023", 100])
                writer.writerow(["30 Nov 2023", 300])
                writer.writerow(["31 Dec 2023", 500])
                for month in range(1, 13):
                    month_end = self._month_end_2024(month)
                    writer.writerow([month_end.strftime("%d %b %Y"), 500 + (10 * month)])

            output = build_method_search_output(
                bank_rate_csv=bank_rate_csv,
                housing_tools_xlsx=housing_tools_xlsx,
                vtuz_csv=vtuz_csv,
                ons_households=1_000_000,
                target_year=2024,
            )

        selected_beta = output.selected_value("BANK_D_INTEREST_D_DEMAND")
        rejected_2024 = next(
            candidate
            for candidate in output.candidates
            if candidate.parameter_key == "BANK_D_INTEREST_D_DEMAND"
            and candidate.candidate_id == "delta_fit_2024_only"
        )
        self.assertGreater(selected_beta, 0.0)
        self.assertLess(rejected_2024.value, 0.0)

    def test_production_cli_writes_expected_output_schema(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            bank_rate_csv = tmp_path / "bank-rate.csv"
            housing_tools_xlsx = tmp_path / "housing-tools.xlsx"
            vtuz_csv = tmp_path / "vtuz.csv"
            output_dir = tmp_path / "evidence"

            bank_rate_csv.write_text(
                "Date Changed,Rate\n"
                "01 Dec 22,2.00\n"
                "01 Jan 24,3.00\n",
                encoding="utf-8",
            )

            workbook = Workbook()
            workbook.remove(workbook.active)
            worksheet = workbook.create_sheet("8. Spreads new mortgage lending")
            for _ in range(4):
                worksheet.append([None, None])
            worksheet.append([date(2023, 12, 31), 3.0])
            for month in range(1, 13):
                month_end = self._month_end_2024(month)
                worksheet.append([month_end, 2.0 + (0.05 * month)])
            workbook.save(housing_tools_xlsx)

            with vtuz_csv.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["DATE", "LPMVTUZ"])
                writer.writerow(["31 Dec 2023", 300])
                for month in range(1, 13):
                    month_end = self._month_end_2024(month)
                    writer.writerow([month_end.strftime("%d %b %Y"), 300 + (20 * month)])

            command = [
                "python3",
                "-m",
                self.calibration_module,
                "--bank-rate-csv",
                str(bank_rate_csv),
                "--housing-tools-xlsx",
                str(housing_tools_xlsx),
                "--vtuz-csv",
                str(vtuz_csv),
                "--ons-households",
                "1000000",
                "--output-dir",
                str(output_dir),
            ]
            result = subprocess.run(
                command,
                cwd=self.repo_root,
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, msg=result.stderr + result.stdout)
            stdout = result.stdout
            for key in (
                "CENTRAL_BANK_INITIAL_BASE_RATE",
                "BANK_INITIAL_RATE",
                "BANK_D_INTEREST_D_DEMAND",
                "BANK_INITIAL_CREDIT_SUPPLY",
            ):
                self.assertIn(f"{key} =", stdout)
            self.assertTrue((output_dir / "BoeBankParameterCalibrationSummary.csv").exists())
            self.assertTrue((output_dir / "BoeBankParameterCalibrationSummary.json").exists())
            self.assertTrue((output_dir / "BoEBankRate2024Monthly.csv").exists())
            self.assertTrue((output_dir / "BoEVTUZSpreadAlignedDeltas2024.csv").exists())
            self.assertTrue((output_dir / "OnsHouseholds2024.csv").exists())

    @staticmethod
    def _month_end_2024(month: int) -> date:
        return date(
            2024,
            month,
            29 if month == 2 else 30 if month in (4, 6, 9, 11) else 31,
        )


if __name__ == "__main__":
    unittest.main()
