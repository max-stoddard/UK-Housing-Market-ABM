from __future__ import annotations

import csv
import json
import subprocess
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from scripts.python.calibration.official.rent_gross_yield_2024 import (
    RENT_GROSS_YIELD_KEY,
    SOURCE_VALUES_FILE_NAME,
    SUMMARY_FILE_NAME,
    build_arg_parser,
    build_calibration_summary,
    extract_hpi_uk_average_prices,
    extract_pipr_rent_values,
    inspect_average_prices_comparator,
    run_calibration,
)


class TestRentGrossYield2024(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.repo_root = Path(__file__).resolve().parents[3]
        cls.module_name = "scripts.python.calibration.official.rent_gross_yield_2024"

    def test_parser_requires_source_paths(self) -> None:
        parser = build_arg_parser()
        args = parser.parse_args(
            [
                "--pipr-xlsx",
                "pipr.xlsx",
                "--hpi-full-csv",
                "hpi.csv",
                "--output-dir",
                "out",
            ]
        )

        self.assertEqual(args.pipr_xlsx, "pipr.xlsx")
        self.assertEqual(args.hpi_full_csv, "hpi.csv")
        self.assertEqual(args.output_dir, "out")
        self.assertIsNone(args.average_prices_csv)

    def test_extract_pipr_rent_values_selects_great_britain_when_uk_unavailable(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            pipr_path = Path(tmp_dir) / "pipr.xlsx"
            self._write_pipr_workbook(pipr_path, rents=[1000 + month for month in range(11)])

            values = extract_pipr_rent_values(pipr_path)

        self.assertEqual(len(values), 11)
        self.assertEqual({item.geography for item in values}, {"Great Britain"})
        self.assertEqual([item.value for item in values[:2]], [1000.0, 1001.0])
        self.assertIn("does not publish UK rent-price levels", values[0].notes)

    def test_extract_pipr_rent_values_rejects_missing_months(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            pipr_path = Path(tmp_dir) / "pipr.xlsx"
            self._write_pipr_workbook(pipr_path, rents=[1000 + month for month in range(10)])

            with self.assertRaisesRegex(ValueError, "Expected PIPR rent-price values for months"):
                extract_pipr_rent_values(pipr_path)

    def test_extract_hpi_uk_average_prices_requires_all_twelve_months(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            hpi_path = Path(tmp_dir) / "hpi.csv"
            self._write_hpi_full_csv(hpi_path, prices=[200_000 + month for month in range(12)])

            values = extract_hpi_uk_average_prices(hpi_path)

        self.assertEqual(len(values), 12)
        self.assertEqual(values[0].date, "2024-01-01")
        self.assertEqual(values[-1].date, "2024-12-01")
        self.assertEqual(values[0].value, 200_000.0)

    def test_average_prices_comparator_is_rejected_without_weight_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            average_prices_path = Path(tmp_dir) / "average-prices.csv"
            self._write_average_prices_csv(average_prices_path)

            comparator = inspect_average_prices_comparator(average_prices_path)

        self.assertEqual(comparator["status"], "rejected_comparator")
        self.assertFalse(comparator["hasWeightFields"])
        self.assertEqual(comparator["targetYearUkRows"], 12)

    def test_build_calibration_summary_contains_expected_value_and_unavailable_uk_note(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            paths = self._build_synthetic_artifacts(Path(tmp_dir))

            summary = build_calibration_summary(**paths)

        selected = summary["selectedConfigValues"][RENT_GROSS_YIELD_KEY]
        self.assertAlmostEqual(selected, round(12 * 1005 / 200_005.5, 10))
        self.assertEqual(summary["calculation"]["rentMean"], 1005.0)
        self.assertEqual(summary["calculation"]["housePriceMean"], 200_005.5)
        self.assertEqual(summary["calculation"]["targetYear"], 2024)
        unavailable = summary["sourceValues"]["unavailablePiprRentLevels"]
        self.assertEqual({item["geography"] for item in unavailable}, {"UK", "Northern Ireland"})
        self.assertEqual(summary["rejectedComparisons"][2]["status"], "rejected_comparator")

    def test_run_calibration_writes_evidence_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_root = Path(tmp_dir)
            paths = self._build_synthetic_artifacts(temp_root)
            output_dir = temp_root / "output"

            summary = run_calibration(output_dir=output_dir, **paths)

            source_values_path = output_dir / SOURCE_VALUES_FILE_NAME
            summary_path = output_dir / SUMMARY_FILE_NAME
            source_values_exists = source_values_path.exists()
            summary_exists = summary_path.exists()
            loaded_summary = json.loads(summary_path.read_text(encoding="utf-8"))

        self.assertTrue(source_values_exists)
        self.assertTrue(summary_exists)
        self.assertEqual(loaded_summary["selectedConfigValues"], summary["selectedConfigValues"])

    def test_cli_writes_expected_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_root = Path(tmp_dir)
            paths = self._build_synthetic_artifacts(temp_root)
            output_dir = temp_root / "output"

            result = subprocess.run(
                [
                    "python3",
                    "-m",
                    self.module_name,
                    "--pipr-xlsx",
                    str(paths["pipr_xlsx"]),
                    "--hpi-full-csv",
                    str(paths["hpi_full_csv"]),
                    "--average-prices-csv",
                    str(paths["average_prices_csv"]),
                    "--output-dir",
                    str(output_dir),
                ],
                cwd=self.repo_root,
                text=True,
                capture_output=True,
                check=False,
            )

        self.assertEqual(result.returncode, 0, msg=result.stderr)
        self.assertIn("RENT_GROSS_YIELD = ", result.stdout)

    @staticmethod
    def _build_synthetic_artifacts(temp_root: Path) -> dict[str, Path]:
        pipr_path = temp_root / "pipr.xlsx"
        hpi_path = temp_root / "hpi.csv"
        average_prices_path = temp_root / "average-prices.csv"
        TestRentGrossYield2024._write_pipr_workbook(pipr_path, rents=[1000 + month for month in range(11)])
        TestRentGrossYield2024._write_hpi_full_csv(hpi_path, prices=[200_000 + month for month in range(12)])
        TestRentGrossYield2024._write_average_prices_csv(average_prices_path)
        return {
            "pipr_xlsx": pipr_path,
            "hpi_full_csv": hpi_path,
            "average_prices_csv": average_prices_path,
        }

    @staticmethod
    def _write_pipr_workbook(path: Path, *, rents: list[float]) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Table 1"
        headers = [
            "Time period",
            "Area code",
            "Area name",
            "Region or country name",
            "Index",
            "Monthly change",
            "Annual change",
            "Rental price",
        ]
        for column, header in enumerate(headers, 1):
            worksheet.cell(row=3, column=column, value=header)

        row_number = 4
        for month, rent in enumerate(rents, 1):
            observed = datetime(2024, month, 1)
            rows = [
                [observed, "K02000001", "UK", "[z]", 100.0, "[x]", "[x]", "[x]"],
                [observed, "K03000001", "Great Britain", "[z]", 100.0, 0.1, 1.0, rent],
                [observed, "N92000002", "Northern Ireland", "[z]", 100.0, "[x]", "[x]", "[x]"],
            ]
            for values in rows:
                for column, value in enumerate(values, 1):
                    worksheet.cell(row=row_number, column=column, value=value)
                row_number += 1
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _write_hpi_full_csv(path: Path, *, prices: list[float]) -> None:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["Date", "RegionName", "AreaCode", "AveragePrice"])
            writer.writeheader()
            for month, price in enumerate(prices, 1):
                writer.writerow(
                    {
                        "Date": f"01/{month:02d}/2024",
                        "RegionName": "United Kingdom",
                        "AreaCode": "K02000001",
                        "AveragePrice": str(price),
                    }
                )

    @staticmethod
    def _write_average_prices_csv(path: Path) -> None:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Date",
                    "Region_Name",
                    "Area_Code",
                    "Average_Price",
                    "Monthly_Change",
                    "Annual_Change",
                    "Average_Price_SA",
                ],
            )
            writer.writeheader()
            for month in range(1, 13):
                writer.writerow(
                    {
                        "Date": f"2024-{month:02d}-01",
                        "Region_Name": "United Kingdom",
                        "Area_Code": "K02000001",
                        "Average_Price": "200000",
                        "Monthly_Change": "",
                        "Annual_Change": "",
                        "Average_Price_SA": "",
                    }
                )


if __name__ == "__main__":
    unittest.main()
