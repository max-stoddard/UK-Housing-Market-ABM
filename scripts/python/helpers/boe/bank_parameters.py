"""Shared helpers for BoE bank-parameter calibration.

@author: Max Stoddard
"""

from __future__ import annotations

import calendar
import csv
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import fmean

from openpyxl import load_workbook


BOE_HOUSING_TOOLS_SHEET = "8. Spreads new mortgage lending"


@dataclass(frozen=True)
class MonthlyObservation:
    """One month-end observation."""

    observation_date: date
    value: float

    @property
    def month_label(self) -> str:
        return self.observation_date.strftime("%Y-%m")


@dataclass(frozen=True)
class CandidateResult:
    """One ranked candidate calibration value."""

    parameter_key: str
    candidate_id: str
    rank: int
    selected: bool
    value: float
    window_label: str
    method_label: str
    rationale: str


@dataclass(frozen=True)
class BankParameterMethodSearchOutput:
    """Resolved source series, diagnostics, and ranked candidates."""

    target_year: int
    ons_households: float
    bank_rate_history: tuple[MonthlyObservation, ...]
    bank_rate_monthly: tuple[MonthlyObservation, ...]
    housing_tools_spread_monthly: tuple[MonthlyObservation, ...]
    housing_tools_spread_target_year: tuple[MonthlyObservation, ...]
    mortgage_rate_proxy_target_year: tuple[MonthlyObservation, ...]
    vtuz_monthly: tuple[MonthlyObservation, ...]
    credit_supply_target_year: tuple[MonthlyObservation, ...]
    delta_panel_full: tuple[dict[str, float | str], ...]
    delta_panel_target_year: tuple[dict[str, float | str], ...]
    candidates: tuple[CandidateResult, ...]

    def selected_value(self, parameter_key: str) -> float:
        for candidate in self.candidates:
            if candidate.parameter_key == parameter_key and candidate.selected:
                return candidate.value
        raise KeyError(f"No selected value for parameter: {parameter_key}")


def _parse_bank_rate_change_date(raw_value: str) -> date:
    return datetime.strptime(raw_value.strip(), "%d %b %y").date()


def load_bank_rate_history(csv_path: Path) -> tuple[tuple[date, float], ...]:
    """Load the BoE Bank Rate change history from CSV."""

    events: list[tuple[date, float]] = []
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            raw_date = row.get("Date Changed", "").strip()
            raw_rate = row.get("Rate", "").strip()
            if not raw_date or not raw_rate:
                continue
            events.append((_parse_bank_rate_change_date(raw_date), float(raw_rate) / 100.0))
    if not events:
        raise ValueError(f"No Bank Rate rows found in: {csv_path}")
    events.sort(key=lambda item: item[0])
    return tuple(events)


def bank_rate_on_day(events: tuple[tuple[date, float], ...], day: date) -> float:
    """Return the active Bank Rate on one day."""

    current_rate: float | None = None
    for changed_on, rate in events:
        if changed_on <= day:
            current_rate = rate
        else:
            break
    if current_rate is None:
        raise ValueError(f"No Bank Rate available for {day.isoformat()}")
    return current_rate


def build_bank_rate_monthly_series(
    events: tuple[tuple[date, float], ...],
    *,
    year: int,
) -> tuple[MonthlyObservation, ...]:
    """Convert change-history Bank Rate data to a daily-weighted monthly series."""

    if year <= 0:
        raise ValueError("year must be positive.")
    monthly_values: list[MonthlyObservation] = []
    for month in range(1, 13):
        month_end_day = calendar.monthrange(year, month)[1]
        month_start = date(year, month, 1)
        month_end = date(year, month, month_end_day)
        day = month_start
        daily_values: list[float] = []
        while day <= month_end:
            daily_values.append(bank_rate_on_day(events, day))
            day += timedelta(days=1)
        monthly_values.append(MonthlyObservation(month_end, fmean(daily_values)))
    return tuple(monthly_values)


def extract_housing_tools_spread_series(workbook_path: Path) -> tuple[MonthlyObservation, ...]:
    """Extract the BoE owner-occupier 2-year 75% LTV spread series from housing-tools."""

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[BOE_HOUSING_TOOLS_SHEET]
        observations: list[MonthlyObservation] = []
        for row in worksheet.iter_rows(min_row=5, values_only=True):
            if not row:
                continue
            stamp = row[0]
            value = row[1]
            if stamp is None or value is None:
                continue
            stamp_date = stamp.date() if hasattr(stamp, "date") else stamp
            if not isinstance(stamp_date, date):
                continue
            observations.append(MonthlyObservation(stamp_date, float(value)))
    finally:
        workbook.close()
    if not observations:
        raise ValueError(f"No spread observations found in: {workbook_path}")
    observations.sort(key=lambda item: item.observation_date)
    return tuple(observations)


def load_vtuz_series(csv_path: Path) -> tuple[MonthlyObservation, ...]:
    """Load the official VTUZ gross-lending series exported from the BoE database."""

    observations: list[MonthlyObservation] = []
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        date_field = reader.fieldnames[0] if reader.fieldnames else "DATE"
        value_field = reader.fieldnames[1] if reader.fieldnames and len(reader.fieldnames) > 1 else "LPMVTUZ"
        for row in reader:
            raw_date = row.get(date_field, "").strip()
            raw_value = row.get(value_field, "").strip()
            if not raw_date or not raw_value:
                continue
            observations.append(
                MonthlyObservation(
                    datetime.strptime(raw_date, "%d %b %Y").date(),
                    float(raw_value),
                )
            )
    if not observations:
        raise ValueError(f"No VTUZ observations found in: {csv_path}")
    observations.sort(key=lambda item: item.observation_date)
    return tuple(observations)


def observations_for_year(
    observations: tuple[MonthlyObservation, ...],
    *,
    year: int,
) -> tuple[MonthlyObservation, ...]:
    """Filter a monthly series to one calendar year."""

    return tuple(item for item in observations if item.observation_date.year == year)


def mortgage_rate_proxy_target_year(
    bank_rate_monthly: tuple[MonthlyObservation, ...],
    spread_target_year: tuple[MonthlyObservation, ...],
) -> tuple[MonthlyObservation, ...]:
    """Combine monthly Bank Rate and housing-tools spread into a mortgage-rate proxy."""

    spread_by_date = {item.observation_date: item.value for item in spread_target_year}
    observations: list[MonthlyObservation] = []
    for bank_rate in bank_rate_monthly:
        spread_pp = spread_by_date.get(bank_rate.observation_date)
        if spread_pp is None:
            raise ValueError(
                f"Missing housing-tools spread for {bank_rate.observation_date.isoformat()}"
            )
        observations.append(
            MonthlyObservation(
                bank_rate.observation_date,
                bank_rate.value + (spread_pp / 100.0),
            )
        )
    return tuple(observations)


def per_household_credit(
    observations: tuple[MonthlyObservation, ...],
    *,
    households: float,
) -> tuple[MonthlyObservation, ...]:
    """Convert gross lending in GBP millions to pounds per household per month."""

    if households <= 0:
        raise ValueError("households must be positive.")
    return tuple(
        MonthlyObservation(item.observation_date, item.value * 1_000_000.0 / households)
        for item in observations
    )


def through_origin_slope(xs: list[float], ys: list[float]) -> float:
    """Fit a through-origin slope."""

    if len(xs) != len(ys):
        raise ValueError("xs and ys must have the same length.")
    if not xs:
        raise ValueError("At least one observation is required.")
    denominator = sum(x * x for x in xs)
    if denominator == 0.0:
        raise ValueError("Cannot fit a through-origin slope with zero x variance.")
    return sum(x * y for x, y in zip(xs, ys)) / denominator


def build_delta_panel(
    spread_observations_pp: tuple[MonthlyObservation, ...],
    credit_observations_per_household: tuple[MonthlyObservation, ...],
) -> tuple[dict[str, float | str], ...]:
    """Align monthly spread and credit series and compute delta diagnostics."""

    spread_by_date = {item.observation_date: item.value / 100.0 for item in spread_observations_pp}
    credit_by_date = {item.observation_date: item.value for item in credit_observations_per_household}
    common_dates = sorted(set(spread_by_date).intersection(credit_by_date))
    if len(common_dates) < 2:
        raise ValueError("At least two overlapping monthly observations are required.")

    rows: list[dict[str, float | str]] = []
    previous_spread = None
    previous_credit = None
    for observation_date in common_dates:
        spread_fraction = spread_by_date[observation_date]
        credit_value = credit_by_date[observation_date]
        delta_spread = None if previous_spread is None else spread_fraction - previous_spread
        delta_credit = None if previous_credit is None else credit_value - previous_credit
        rows.append(
            {
                "month_end": observation_date.isoformat(),
                "spread_fraction": spread_fraction,
                "credit_per_household": credit_value,
                "delta_spread_fraction": delta_spread,
                "delta_credit_per_household": delta_credit,
            }
        )
        previous_spread = spread_fraction
        previous_credit = credit_value
    return tuple(rows)


def _slope_from_delta_panel(delta_panel: tuple[dict[str, float | str], ...]) -> float:
    xs = [
        float(row["delta_credit_per_household"])
        for row in delta_panel
        if row["delta_credit_per_household"] is not None
    ]
    ys = [
        float(row["delta_spread_fraction"])
        for row in delta_panel
        if row["delta_spread_fraction"] is not None
    ]
    return through_origin_slope(xs, ys)


def build_method_search_output(
    *,
    bank_rate_csv: Path,
    housing_tools_xlsx: Path,
    vtuz_csv: Path,
    ons_households: float,
    target_year: int,
) -> BankParameterMethodSearchOutput:
    """Resolve the agreed BoE bank-parameter method search output."""

    bank_rate_events = load_bank_rate_history(bank_rate_csv)
    bank_rate_monthly = build_bank_rate_monthly_series(bank_rate_events, year=target_year)
    bank_rate_history = tuple(
        MonthlyObservation(changed_on, rate)
        for changed_on, rate in bank_rate_events
    )
    housing_tools_spread_monthly = extract_housing_tools_spread_series(housing_tools_xlsx)
    housing_tools_spread_target_year = observations_for_year(
        housing_tools_spread_monthly,
        year=target_year,
    )
    vtuz_monthly = load_vtuz_series(vtuz_csv)
    credit_supply_monthly = per_household_credit(vtuz_monthly, households=ons_households)
    credit_supply_target_year = observations_for_year(
        credit_supply_monthly,
        year=target_year,
    )
    mortgage_rate_proxy = mortgage_rate_proxy_target_year(
        bank_rate_monthly,
        housing_tools_spread_target_year,
    )

    delta_panel_full = build_delta_panel(
        tuple(
            item
            for item in housing_tools_spread_monthly
            if item.observation_date <= date(target_year, 12, 31)
        ),
        tuple(
            item
            for item in credit_supply_monthly
            if item.observation_date <= date(target_year, 12, 31)
        ),
    )
    delta_panel_target_year = build_delta_panel(
        housing_tools_spread_target_year,
        credit_supply_target_year,
    )

    selected_base_rate = fmean(item.value for item in bank_rate_monthly)
    selected_initial_rate = fmean(item.value for item in mortgage_rate_proxy)
    selected_credit_supply = fmean(item.value for item in credit_supply_target_year)
    selected_beta = _slope_from_delta_panel(delta_panel_full)
    beta_target_year = _slope_from_delta_panel(delta_panel_target_year)

    candidates = (
        CandidateResult(
            parameter_key="CENTRAL_BANK_INITIAL_BASE_RATE",
            candidate_id="full_year_2024_mean",
            rank=1,
            selected=True,
            value=selected_base_rate,
            window_label=f"{target_year} Jan-Dec",
            method_label="Daily-weighted monthly Bank Rate mean",
            rationale="Selected default: full-year 2024 constant proxy aligned to the validation window.",
        ),
        CandidateResult(
            parameter_key="CENTRAL_BANK_INITIAL_BASE_RATE",
            candidate_id="dec_2024_snapshot",
            rank=2,
            selected=False,
            value=bank_rate_monthly[-1].value,
            window_label=f"{target_year}-12",
            method_label="December month-end snapshot",
            rationale="Rejected: year-end snapshot is less aligned to the full-year validation window.",
        ),
        CandidateResult(
            parameter_key="CENTRAL_BANK_INITIAL_BASE_RATE",
            candidate_id="jan_2024_snapshot",
            rank=3,
            selected=False,
            value=bank_rate_monthly[0].value,
            window_label=f"{target_year}-01",
            method_label="January month-end snapshot",
            rationale="Rejected: opening-month snapshot is less representative than the full-year mean.",
        ),
        CandidateResult(
            parameter_key="BANK_INITIAL_RATE",
            candidate_id="full_year_2024_mean",
            rank=1,
            selected=True,
            value=selected_initial_rate,
            window_label=f"{target_year} Jan-Dec",
            method_label="Monthly Bank Rate plus monthly housing-tools spread",
            rationale="Selected default: full-year 2024 mortgage-rate proxy aligned to the validation window.",
        ),
        CandidateResult(
            parameter_key="BANK_INITIAL_RATE",
            candidate_id="dec_2024_snapshot",
            rank=2,
            selected=False,
            value=mortgage_rate_proxy[-1].value,
            window_label=f"{target_year}-12",
            method_label="December mortgage-rate proxy snapshot",
            rationale="Rejected: year-end snapshot is less aligned to the full-year validation window.",
        ),
        CandidateResult(
            parameter_key="BANK_INITIAL_RATE",
            candidate_id="jan_2024_snapshot",
            rank=3,
            selected=False,
            value=mortgage_rate_proxy[0].value,
            window_label=f"{target_year}-01",
            method_label="January mortgage-rate proxy snapshot",
            rationale="Rejected: opening-month snapshot is less representative than the full-year mean.",
        ),
        CandidateResult(
            parameter_key="BANK_INITIAL_CREDIT_SUPPLY",
            candidate_id="full_year_2024_mean",
            rank=1,
            selected=True,
            value=selected_credit_supply,
            window_label=f"{target_year} Jan-Dec",
            method_label="Monthly VTUZ converted to pounds per household",
            rationale="Selected default: full-year 2024 per-household lending mean aligned to the validation window.",
        ),
        CandidateResult(
            parameter_key="BANK_INITIAL_CREDIT_SUPPLY",
            candidate_id="dec_2024_snapshot",
            rank=2,
            selected=False,
            value=credit_supply_target_year[-1].value,
            window_label=f"{target_year}-12",
            method_label="December VTUZ per-household snapshot",
            rationale="Rejected: year-end snapshot is less aligned to the full-year validation window.",
        ),
        CandidateResult(
            parameter_key="BANK_INITIAL_CREDIT_SUPPLY",
            candidate_id="jan_2024_snapshot",
            rank=3,
            selected=False,
            value=credit_supply_target_year[0].value,
            window_label=f"{target_year}-01",
            method_label="January VTUZ per-household snapshot",
            rationale="Rejected: opening-month snapshot is less representative than the full-year mean.",
        ),
        CandidateResult(
            parameter_key="BANK_D_INTEREST_D_DEMAND",
            candidate_id="delta_fit_1995_to_2024",
            rank=1,
            selected=True,
            value=selected_beta,
            window_label=f"{housing_tools_spread_monthly[0].observation_date.year}-{target_year}",
            method_label="Through-origin fit on monthly spread deltas versus monthly credit deltas",
            rationale="Selected default: full pre-2025 overlap stays positive and matches the model update equation.",
        ),
        CandidateResult(
            parameter_key="BANK_D_INTEREST_D_DEMAND",
            candidate_id="delta_fit_2024_only",
            rank=2,
            selected=False,
            value=beta_target_year,
            window_label=f"{target_year} Jan-Dec",
            method_label="Through-origin fit on monthly spread deltas versus monthly credit deltas",
            rationale="Rejected diagnostic: 2024-only window turns the fitted coefficient negative.",
        ),
    )

    return BankParameterMethodSearchOutput(
        target_year=target_year,
        ons_households=ons_households,
        bank_rate_history=bank_rate_history,
        bank_rate_monthly=bank_rate_monthly,
        housing_tools_spread_monthly=housing_tools_spread_monthly,
        housing_tools_spread_target_year=housing_tools_spread_target_year,
        mortgage_rate_proxy_target_year=mortgage_rate_proxy,
        vtuz_monthly=vtuz_monthly,
        credit_supply_target_year=credit_supply_target_year,
        delta_panel_full=delta_panel_full,
        delta_panel_target_year=delta_panel_target_year,
        candidates=candidates,
    )
