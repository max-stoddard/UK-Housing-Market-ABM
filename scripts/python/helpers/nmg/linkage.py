#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Respondent-linkage helpers for NMG expectation experiments.

@author: Max Stoddard
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from scripts.python.helpers.nmg.hpa_expectation import NmgWaveData

_LINKAGE_SHEET_NAME = "2011-2025 PID-SUBSID"


@dataclass(frozen=True)
class NmgLinkageRecord:
    wave_label: str
    subsid: str | None
    pid: str | None
    canonical_id: str | None


@dataclass(frozen=True)
class NmgRespondentLinkage:
    records_by_wave_label: dict[str, tuple[NmgLinkageRecord, ...]]
    subsid_to_canonical_by_wave_label: dict[str, dict[str, str]]
    pid_to_canonical_by_wave_label: dict[str, dict[str, str]]


def _normalize_identifier(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "none":
        return None
    if re.fullmatch(r"-?\d+\.0", text):
        return text[:-2]
    return text


def normalize_linkage_wave_label(raw_wave_label: object) -> str:
    text = str(raw_wave_label).strip()
    if "March 2025" in text:
        return "2025-pt1"
    if "September 2025" in text:
        return "2025-pt2"
    year_match = re.search(r"(20\d{2}|201\d)", text)
    if year_match is None:
        raise ValueError(f"Could not normalize linkage wave label: {text}")
    return year_match.group(1)


def load_pid_subsid_linkage(path: Path) -> NmgRespondentLinkage:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if _LINKAGE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook is missing linkage sheet {_LINKAGE_SHEET_NAME!r}: {path}")
    worksheet = workbook[_LINKAGE_SHEET_NAME]
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    column_lookup = {str(value): index for index, value in enumerate(header) if value is not None}
    required_columns = {"wave", "subsidn", "pid"}
    missing_columns = required_columns.difference(column_lookup)
    if missing_columns:
        missing_text = ", ".join(sorted(missing_columns))
        raise ValueError(f"Linkage sheet is missing required columns: {missing_text}")

    records_by_wave_label: dict[str, list[NmgLinkageRecord]] = {}
    subsid_to_canonical_by_wave_label: dict[str, dict[str, str]] = {}
    pid_to_canonical_by_wave_label: dict[str, dict[str, str]] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        wave_label = normalize_linkage_wave_label(row[column_lookup["wave"]])
        subsid = _normalize_identifier(row[column_lookup["subsidn"]])
        pid = _normalize_identifier(row[column_lookup["pid"]])
        canonical_id = None
        if pid is not None:
            canonical_id = f"pid:{pid}"
        elif subsid is not None:
            canonical_id = f"subsid:{subsid}"
        record = NmgLinkageRecord(
            wave_label=wave_label,
            subsid=subsid,
            pid=pid,
            canonical_id=canonical_id,
        )
        records_by_wave_label.setdefault(wave_label, []).append(record)
        if canonical_id is None:
            continue
        if subsid is not None:
            subsid_to_canonical_by_wave_label.setdefault(wave_label, {})[subsid] = canonical_id
        if pid is not None:
            pid_to_canonical_by_wave_label.setdefault(wave_label, {})[pid] = canonical_id

    linkage = NmgRespondentLinkage(
        records_by_wave_label={key: tuple(value) for key, value in records_by_wave_label.items()},
        subsid_to_canonical_by_wave_label=subsid_to_canonical_by_wave_label,
        pid_to_canonical_by_wave_label=pid_to_canonical_by_wave_label,
    )
    workbook.close()
    return linkage


def resolve_wave_row_canonical_id(
    linkage: NmgRespondentLinkage,
    *,
    wave_label: str,
    row: dict[str, str],
) -> str | None:
    pid_lookup = linkage.pid_to_canonical_by_wave_label.get(wave_label, {})
    subsid_lookup = linkage.subsid_to_canonical_by_wave_label.get(wave_label, {})

    pid = _normalize_identifier(row.get("pid"))
    if pid is not None:
        return pid_lookup.get(pid, f"pid:{pid}")

    subsid = _normalize_identifier(row.get("subsid"))
    if subsid is not None:
        return subsid_lookup.get(subsid, f"subsid:{subsid}")
    return None


def build_matched_panel_row_indices(
    waves: dict[str, NmgWaveData],
    *,
    required_wave_labels: tuple[str, ...],
    linkage: NmgRespondentLinkage,
) -> dict[str, set[int]]:
    if not required_wave_labels:
        raise ValueError("required_wave_labels cannot be empty.")

    canonical_ids_by_wave_label: dict[str, dict[str, int]] = {}
    for wave_label in required_wave_labels:
        try:
            wave = waves[wave_label]
        except KeyError as exc:
            raise ValueError(f"Missing NMG wave required for matched-panel construction: {wave_label}") from exc
        canonical_ids_for_wave: dict[str, int] = {}
        for row_index, row in enumerate(wave.rows):
            canonical_id = resolve_wave_row_canonical_id(
                linkage,
                wave_label=wave_label,
                row=row,
            )
            if canonical_id is None or canonical_id in canonical_ids_for_wave:
                continue
            canonical_ids_for_wave[canonical_id] = row_index
        if not canonical_ids_for_wave:
            raise ValueError(f"No matched-panel identifiers were found for wave {wave_label}.")
        canonical_ids_by_wave_label[wave_label] = canonical_ids_for_wave

    shared_canonical_ids = set.intersection(
        *(set(canonical_ids_by_wave_label[wave_label]) for wave_label in required_wave_labels)
    )
    if not shared_canonical_ids:
        raise ValueError("No shared respondent identifiers were found across the requested waves.")

    return {
        wave_label: {
            canonical_ids_by_wave_label[wave_label][canonical_id]
            for canonical_id in shared_canonical_ids
        }
        for wave_label in required_wave_labels
    }


__all__ = [
    "NmgLinkageRecord",
    "NmgRespondentLinkage",
    "build_matched_panel_row_indices",
    "load_pid_subsid_linkage",
    "normalize_linkage_wave_label",
    "resolve_wave_row_canonical_id",
]
