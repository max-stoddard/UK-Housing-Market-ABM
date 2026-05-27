"""Review and reproducibility checks for the 2024 validation catalog.

@author: Max Stoddard
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from statistics import fmean
from typing import Any

from openpyxl import load_workbook

from scripts.python.validation.model.extractors import HOUSEHOLD_DISTRIBUTION_SPECS
from scripts.python.validation.model.hpi import (
    compute_rebased_mean,
    compute_rebased_std,
    estimate_dominant_cycle_period_months,
    load_hmlr_uk_full_file_series,
)
from scripts.python.validation.model.validation_catalog_2024 import (
    ADVANCES_TARGET_TOLERANCE,
    BOE_HOUSING_TOOLS_2024_COMPARISON_SCALE_BY_METRIC_ID,
    FPC_SOURCE_2024_BY_METRIC_ID,
    HPI_2024_CYCLE_PERIOD_MONTHS,
    HPI_2024_REBASED_MEAN,
    HPI_FULL_HISTORY_REBASED_STD,
    HPI_TARGET_TOLERANCE,
    HOUSEHOLD_OWNING_SHARE_2024,
    HOUSEHOLD_RENTING_SHARE_2024,
    INTEREST_RATE_SPREAD_2024_QUARTERLY_MEANS,
    MARKET_SOURCE_2024_BY_METRIC_ID,
    OO_DEBT_TO_INCOME_2024_QUARTERLY_VALUES,
    RENTAL_YIELD_2024_QUARTERLY_VALUES,
    RPI_2024_GB_REBASED_MEAN,
    RPI_2024_GB_REBASED_MONTHLY_VALUES,
    SUPPORTED_FPC_METRIC_IDS,
    TARGETS_BY_ID,
    UKF_SOURCE_2024_BY_METRIC_ID,
    UNSUPPORTED_FPC_METRIC_IDS,
)

REVIEW_LEDGER_PATH = "scripts/python/validation/model/validation_catalog_2024_review.json"
ONS_QWND_SNAPSHOT_PATH = "input-data-versions/validation-sources/2024/ons/qwnd-household-gross-disposable-income-2023q2-2024q4.json"
HMLR_HPI_SOURCE_PATH = "input-data-versions/validation-sources/2024/hmlr/UK-HPI-full-file-2024-12.csv"
FRS_TENURE_2024_SOURCE_PATH = "input-data-versions/validation-sources/2024/frs/frs-2023-24-tenure-tables.xlsx"
ONS_RPI_2024_SOURCE_PATH = "input-data-versions/validation-sources/2024/ons-rpi/priceindexofprivaterentsukhistoricalseries-2025-03-26.xlsx"


@dataclass(frozen=True)
class CatalogReviewResult:
    """Structured result for the catalog audit."""

    success: bool
    errors: list[str]
    live_review_data: dict[str, object]


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[4]


def _band(lower: float, upper: float) -> dict[str, float]:
    return {"lower": float(lower), "upper": float(upper)}


def _round_three_dp(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))


def _serialize_core_metric(metric_id: str) -> dict[str, object]:
    metric = TARGETS_BY_ID[metric_id]
    return {
        "requirement": metric.requirement,
        "units": metric.units,
        "source_label": metric.source_label,
        "target_band": _band(metric.target_band.lower, metric.target_band.upper),
        "file_name": metric.file_name,
        "scale": metric.scale,
    }


def _serialize_household_metric(metric_id: str) -> dict[str, object]:
    metric = TARGETS_BY_ID[metric_id]
    return {
        "requirement": metric.requirement,
        "units": metric.units,
        "source_label": metric.source_label,
        "target_band": _band(metric.target_band.lower, metric.target_band.upper),
        "legacy_validation_module": metric.legacy_validation_module,
        "results_file_name": HOUSEHOLD_DISTRIBUTION_SPECS[metric_id].results_file_name,
    }


def _serialize_household_share_metric(metric_id: str) -> dict[str, object]:
    metric = TARGETS_BY_ID[metric_id]
    return {
        "requirement": metric.requirement,
        "units": metric.units,
        "source_label": metric.source_label,
        "target_band": _band(metric.target_band.lower, metric.target_band.upper),
        "file_name": metric.file_name,
        "loss_family": metric.loss_family,
    }


def _scored_metric_ids() -> list[str]:
    return [
        metric_id
        for metric_id, metric in TARGETS_BY_ID.items()
        if metric.requirement == "required" and metric.target_band is not None
    ]


def _extract_single(pattern: str, text: str, *, cast: type = float) -> Any:
    match = re.search(pattern, text, flags=re.MULTILINE | re.DOTALL)
    if match is None:
        raise RuntimeError(f"Unable to extract required value with pattern: {pattern}")
    raw = match.group(1)
    return cast(raw)


def extract_fpc_latest_values(text_path: Path) -> dict[str, float]:
    text = text_path.read_text(encoding="utf-8")
    return {
        "core_mortgageApprovals": float(
            _extract_single(r"Conditions and terms in markets\s*([0-9]+)\s*4:\s*Mortgage\s*approvals", text, cast=int)
        ),
        "core_housingTransactions": float(
            _extract_single(r"90050\s*([0-9]+)\s*\(Mar 2024\)\s*6:\s*House price", text, cast=int)
        ),
        "core_debtToIncome": _extract_single(
            r"0\.7%\s*\(2023Q4\)\s*([0-9]+\.[0-9]+)%\s*\(2023Q4\)",
            text,
        ),
        "core_housePriceGrowth": _extract_single(
            r"-1\.2%\s*([0-9]+\.[0-9]+)%\s*\(Mar 2024\)\s*7:\s*House price to",
            text,
        ),
        "core_priceToIncome": _extract_single(
            r"5\.9\s*([0-9]+\.[0-9]+)\s*\(2023Q4\)\s*8:\s*Spreads on new",
            text,
        ),
        "core_interestRateSpread": _extract_single(
            r"51048\s*\(Mar 2024\)\s*([0-9]+\.[0-9]+)\s*\(Mar 2024\)",
            text,
        ),
    }


def extract_ukf_advances_from_evidence(text_path: Path) -> dict[str, object]:
    text = text_path.read_text(encoding="utf-8")
    ftb = float(_extract_single(r"First-time buyers:\s*([0-9,]+)", text, cast=str).replace(",", ""))
    hm = float(_extract_single(r"Homemovers:\s*([0-9,]+)", text, cast=str).replace(",", ""))
    return {
        "core_advancesToFTB": {"raw_source_value": ftb, "normalized_source_value": ftb / 12.0 / 1_000.0},
        "core_advancesToHM": {"raw_source_value": hm, "normalized_source_value": hm / 12.0 / 1_000.0},
    }


def extract_ukf_btl_advances_from_evidence(text_path: Path) -> dict[str, object]:
    text = text_path.read_text(encoding="utf-8")
    quarter_values = [
        float(_extract_single(r"Q1 2024 House purchase:\s*([0-9,]+)", text, cast=str).replace(",", "")),
        float(_extract_single(r"Q2 2024 House purchase:\s*([0-9,]+)", text, cast=str).replace(",", "")),
        float(_extract_single(r"Q3 2024 House purchase:\s*([0-9,]+)", text, cast=str).replace(",", "")),
        float(_extract_single(r"Q4 2024 House purchase:\s*([0-9,]+)", text, cast=str).replace(",", "")),
    ]
    annual_total = float(sum(quarter_values))
    return {
        "raw_source_value": annual_total,
        "normalized_source_value": annual_total / 12.0 / 1_000.0,
        "source_reference_raw_values": quarter_values,
    }


def extract_rental_yield_from_evidence(text_path: Path) -> list[float]:
    text = text_path.read_text(encoding="utf-8")
    return [
        _extract_single(r"Q1 2024:\s*([0-9]+\.[0-9]+)", text),
        _extract_single(r"Q2 2024:\s*([0-9]+\.[0-9]+)", text),
        _extract_single(r"Q3 2024:\s*([0-9]+\.[0-9]+)", text),
        _extract_single(r"Q4 2024:\s*([0-9]+\.[0-9]+)", text),
    ]


def _coerce_workbook_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, date_format)
            except ValueError:
                continue
    return None


def extract_boe_housing_tools_values_2024(
    workbook_path: Path,
    *,
    sheet_name: str,
    expected_count: int,
) -> list[float]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        values: list[float] = []
        for row in worksheet.iter_rows(min_row=5, values_only=True):
            if not row:
                continue
            timestamp = _coerce_workbook_datetime(row[0])
            value = row[1]
            if timestamp is not None and timestamp.year == 2024 and value is not None:
                values.append(float(value))
        if len(values) != expected_count:
            raise RuntimeError(f"Expected {expected_count} 2024 values from {sheet_name}, found {len(values)}")
        return values
    finally:
        workbook.close()


def extract_spread_monthly_values_2024(workbook_path: Path) -> list[float]:
    return extract_boe_housing_tools_values_2024(
        workbook_path,
        sheet_name="8. Spreads new mortgage lending",
        expected_count=12,
    )


def extract_frs_tenure_values_2024(workbook_path: Path) -> dict[str, float]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["3_6"]
        header = [cell for cell in next(worksheet.iter_rows(min_row=9, max_row=9, values_only=True))]
        columns = {str(label): index for index, label in enumerate(header) if label is not None}
        for row in worksheet.iter_rows(min_row=10, values_only=True):
            if row[0] == "2023/24":
                owned_outright = float(row[columns["Owned outright"]])
                buying_with_mortgage = float(row[columns["Buying with a mortgage [Note 14]"]])
                private_renting = float(row[columns["Private renting sector [Note 2]"]])
                return {
                    "household_owning_share": owned_outright + buying_with_mortgage,
                    "household_renting_share": private_renting,
                }
        raise RuntimeError("Unable to locate 2023/24 row in FRS tenure table 3_6")
    finally:
        workbook.close()


def extract_rpi_rebased_values(workbook_path: Path, *, year: int, geography: str = "Great Britain") -> list[float]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["Table 1"]
        header = [cell for cell in next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))]
        try:
            geography_column = header.index(geography)
        except ValueError as exc:
            raise RuntimeError(f"Unable to locate geography {geography!r} in ONS RPI table") from exc
        values: list[float] = []
        for row in worksheet.iter_rows(min_row=5, values_only=True):
            timestamp = row[0]
            if isinstance(timestamp, datetime) and timestamp.year == year:
                values.append(float(row[geography_column]))
        if len(values) != 12:
            raise RuntimeError(f"Expected 12 monthly {geography} RPI values for {year}, found {len(values)}")
        first = values[0]
        if first <= 0.0:
            raise RuntimeError(f"Cannot rebase non-positive {geography} RPI value for {year}")
        return [value / first for value in values]
    finally:
        workbook.close()


def _quarter_means(values: list[float]) -> list[float]:
    if len(values) % 3 != 0:
        raise ValueError("Quarter mean helper expects a multiple of three monthly values")
    return [float(fmean(values[index : index + 3])) for index in range(0, len(values), 3)]


def _scaled_boe_housing_tools_value(metric_id: str, raw_value: float) -> float:
    scale = BOE_HOUSING_TOOLS_2024_COMPARISON_SCALE_BY_METRIC_ID[metric_id]
    if scale == 0.001:
        return raw_value / 1_000.0
    return raw_value * scale


def _find_quarter_column(worksheet, *, year_row: int, quarter_row: int, year: int, quarter: str) -> int:
    active_year: int | None = None
    for column_index in range(1, worksheet.max_column + 1):
        year_value = worksheet.cell(row=year_row, column=column_index).value
        if isinstance(year_value, int):
            active_year = year_value
        quarter_value = worksheet.cell(row=quarter_row, column=column_index).value
        if active_year == year and quarter_value == quarter:
            return column_index
    raise RuntimeError(f"Unable to locate {year} {quarter} in worksheet {worksheet.title}")


def extract_oo_dti_source_components_2024(workbook_path: Path) -> dict[str, dict[str, float]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet_111 = workbook["1.11"]
        sheet_133 = workbook["1.33"]
        components: dict[str, dict[str, float]] = {}
        for quarter_label in ("Q1", "Q2", "Q3", "Q4"):
            column_111 = _find_quarter_column(sheet_111, year_row=13, quarter_row=14, year=2024, quarter=quarter_label)
            column_133 = _find_quarter_column(sheet_133, year_row=12, quarter_row=13, year=2024, quarter=quarter_label)
            key = f"2024{quarter_label}"
            components[key] = {
                "aggregate_debt": float(sheet_111.cell(row=33, column=column_111).value),
                "regulated_total": float(sheet_133.cell(row=53, column=column_133).value),
                "regulated_btl_share": float(sheet_133.cell(row=41, column=column_133).value),
                "nonregulated_total": float(sheet_133.cell(row=95, column=column_133).value),
                "nonregulated_btl_share": float(sheet_133.cell(row=91, column=column_133).value),
            }
        return components
    finally:
        workbook.close()


def load_ons_qwnd_snapshot(snapshot_path: Path) -> dict[str, float]:
    payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
    quarterly_values = payload["quarterlyValues"]
    return {quarter: float(value) for quarter, value in quarterly_values.items()}


def _compute_trailing_four_quarter_qwnd(quarterly_values: dict[str, float]) -> dict[str, float]:
    q1 = quarterly_values["2023Q2"] + quarterly_values["2023Q3"] + quarterly_values["2023Q4"] + quarterly_values["2024Q1"]
    q2 = quarterly_values["2023Q3"] + quarterly_values["2023Q4"] + quarterly_values["2024Q1"] + quarterly_values["2024Q2"]
    q3 = quarterly_values["2023Q4"] + quarterly_values["2024Q1"] + quarterly_values["2024Q2"] + quarterly_values["2024Q3"]
    q4 = quarterly_values["2024Q1"] + quarterly_values["2024Q2"] + quarterly_values["2024Q3"] + quarterly_values["2024Q4"]
    return {"2024Q1": q1, "2024Q2": q2, "2024Q3": q3, "2024Q4": q4}


def _compute_oo_dti_quarterly_values(
    source_components: dict[str, dict[str, float]],
    trailing_four_quarter_qwnd: dict[str, float],
) -> list[float]:
    values: list[float] = []
    for quarter in ("2024Q1", "2024Q2", "2024Q3", "2024Q4"):
        components = source_components[quarter]
        btl_unsecuritised = (
            components["regulated_total"] * components["regulated_btl_share"] / 100.0
            + components["nonregulated_total"] * components["nonregulated_btl_share"] / 100.0
        )
        oo_share = 1.0 - (btl_unsecuritised / (components["regulated_total"] + components["nonregulated_total"]))
        oo_balance = components["aggregate_debt"] * oo_share
        values.append(100.0 * oo_balance / trailing_four_quarter_qwnd[quarter])
    return values


def build_live_review_data(repo_root: Path | None = None) -> dict[str, object]:
    repo_root = repo_root or _repo_root()

    fpc_values = extract_fpc_latest_values(repo_root / "input-data-versions/validation-sources/2024/cis/fpc-core-indicators-june-2024.txt")
    ukf_values = extract_ukf_advances_from_evidence(
        repo_root / "input-data-versions/validation-sources/2024/ukf/household-finance-review-2024-q4-validation-evidence.txt"
    )
    ukf_btl_values = extract_ukf_btl_advances_from_evidence(
        repo_root / "input-data-versions/validation-sources/2024/ukf/btl-mortgage-market-update-2024-validation-evidence.txt"
    )
    rental_yield_values = extract_rental_yield_from_evidence(
        repo_root / "input-data-versions/validation-sources/2024/ukf/btl-rental-yield-2024-validation-evidence.txt"
    )
    boe_housing_tools_workbook = repo_root / "input-data-versions/validation-sources/2024/boe/housing-tools.xlsx"
    boe_housing_tools_core_values = {
        "core_mortgageApprovals": extract_boe_housing_tools_values_2024(
            boe_housing_tools_workbook,
            sheet_name="4.Mortgage approvals",
            expected_count=12,
        ),
        "core_housingTransactions": extract_boe_housing_tools_values_2024(
            boe_housing_tools_workbook,
            sheet_name="5.Housing transactions",
            expected_count=12,
        ),
        "core_debtToIncome": extract_boe_housing_tools_values_2024(
            boe_housing_tools_workbook,
            sheet_name="3. Household debt to income",
            expected_count=4,
        ),
        "core_housePriceGrowth": extract_boe_housing_tools_values_2024(
            boe_housing_tools_workbook,
            sheet_name="6.House price growth",
            expected_count=12,
        ),
        "core_priceToIncome": extract_boe_housing_tools_values_2024(
            boe_housing_tools_workbook,
            sheet_name="7.House prices disp. income",
            expected_count=4,
        ),
    }
    spread_monthly_values = extract_spread_monthly_values_2024(boe_housing_tools_workbook)
    spread_quarterly_values = _quarter_means(spread_monthly_values)
    tenure_values = extract_frs_tenure_values_2024(repo_root / FRS_TENURE_2024_SOURCE_PATH)
    rpi_rebased_values = extract_rpi_rebased_values(repo_root / ONS_RPI_2024_SOURCE_PATH, year=2024)
    oo_source_components = extract_oo_dti_source_components_2024(repo_root / "input-data-versions/validation-sources/2024/mlar/mlar-longrun-detailed.xlsx")
    ons_qwnd_quarterly_values = load_ons_qwnd_snapshot(repo_root / ONS_QWND_SNAPSHOT_PATH)
    trailing_four_quarter_qwnd = _compute_trailing_four_quarter_qwnd(ons_qwnd_quarterly_values)
    oo_dti_quarterly_values = _compute_oo_dti_quarterly_values(oo_source_components, trailing_four_quarter_qwnd)
    hpi_2024_sa_values = load_hmlr_uk_full_file_series(
        repo_root / HMLR_HPI_SOURCE_PATH,
        field_name="IndexSA",
        start_year_month=(2024, 1),
        end_year_month=(2024, 12),
    )
    hpi_full_sa_values = load_hmlr_uk_full_file_series(
        repo_root / HMLR_HPI_SOURCE_PATH,
        field_name="IndexSA",
    )
    rebased_hpi_2024_values = [float(value) for value in load_hmlr_uk_full_file_series(
        repo_root / HMLR_HPI_SOURCE_PATH,
        field_name="IndexSA",
        start_year_month=(2024, 1),
        end_year_month=(2024, 12),
    )]
    rebased_hpi_2024_values = [value / rebased_hpi_2024_values[0] for value in rebased_hpi_2024_values]
    hpi_full_index_values = load_hmlr_uk_full_file_series(repo_root / HMLR_HPI_SOURCE_PATH, field_name="Index")

    source_fpc_core_metrics = {
        metric_id: {
            "raw_source_value": value,
            "normalized_source_value": value,
            "source_as_of": FPC_SOURCE_2024_BY_METRIC_ID[metric_id].source_as_of,
            "source_indicator_label": FPC_SOURCE_2024_BY_METRIC_ID[metric_id].source_indicator_label,
            "mapping_status": FPC_SOURCE_2024_BY_METRIC_ID[metric_id].mapping_status,
        }
        for metric_id, value in fpc_values.items()
    }
    source_fpc_core_metrics["core_mortgageApprovals"]["normalized_source_value"] = fpc_values["core_mortgageApprovals"] / 1_000.0
    source_fpc_core_metrics["core_housingTransactions"]["normalized_source_value"] = (
        fpc_values["core_housingTransactions"] / 1_000.0
    )

    source_ukf_advances_metrics = {
        "core_advancesToFTB": {
            **ukf_values["core_advancesToFTB"],
            "target_band": _band(
                _round_three_dp(ukf_values["core_advancesToFTB"]["normalized_source_value"] * (1.0 - ADVANCES_TARGET_TOLERANCE)),
                _round_three_dp(ukf_values["core_advancesToFTB"]["normalized_source_value"] * (1.0 + ADVANCES_TARGET_TOLERANCE)),
            ),
            "source_reference_raw_values": [ukf_values["core_advancesToFTB"]["raw_source_value"]],
        },
        "core_advancesToHM": {
            **ukf_values["core_advancesToHM"],
            "target_band": _band(
                _round_three_dp(ukf_values["core_advancesToHM"]["normalized_source_value"] * (1.0 - ADVANCES_TARGET_TOLERANCE)),
                _round_three_dp(ukf_values["core_advancesToHM"]["normalized_source_value"] * (1.0 + ADVANCES_TARGET_TOLERANCE)),
            ),
            "source_reference_raw_values": [ukf_values["core_advancesToHM"]["raw_source_value"]],
        },
        "core_advancesToBTL": {
            **ukf_btl_values,
            "target_band": _band(
                _round_three_dp(ukf_btl_values["normalized_source_value"] * (1.0 - ADVANCES_TARGET_TOLERANCE)),
                _round_three_dp(ukf_btl_values["normalized_source_value"] * (1.0 + ADVANCES_TARGET_TOLERANCE)),
            ),
        },
    }

    source_boe_housing_tools_core_metrics = {}
    for metric_id, raw_values in boe_housing_tools_core_values.items():
        comparison_values = [_scaled_boe_housing_tools_value(metric_id, value) for value in raw_values]
        source_boe_housing_tools_core_metrics[metric_id] = {
            "raw_values": raw_values,
            "value_count": len(raw_values),
            "raw_source_value": float(fmean(raw_values)),
            "normalized_source_value": _scaled_boe_housing_tools_value(metric_id, float(fmean(raw_values))),
            "target_band": _band(min(comparison_values), max(comparison_values)),
        }

    return {
        "supported_fpc_metric_ids": list(SUPPORTED_FPC_METRIC_IDS),
        "unsupported_fpc_metric_ids": list(UNSUPPORTED_FPC_METRIC_IDS),
        "advances_target_tolerance": ADVANCES_TARGET_TOLERANCE,
        "metric_weighting_and_composite_aggregation": {
            "scored_metric_ids": _scored_metric_ids(),
            "required_metric_count": len(_scored_metric_ids()),
            "metric_weight": 1.0,
            "composite_rule": "weighted_mean",
        },
        "source_fpc_core_metrics": source_fpc_core_metrics,
        "source_ukf_advances_metrics": source_ukf_advances_metrics,
        "source_boe_housing_tools_core_metrics": source_boe_housing_tools_core_metrics,
        "source_market_interest_rate_spread": {
            "quarterly_values": spread_quarterly_values,
            "annual_mean": float(fmean(spread_quarterly_values)),
            "target_band": _band(min(spread_quarterly_values), max(spread_quarterly_values)),
        },
        "source_market_rental_yield": {
            "quarterly_values": rental_yield_values,
            "annual_mean": float(fmean(rental_yield_values)),
            "target_band": _band(min(rental_yield_values), max(rental_yield_values)),
        },
        "source_market_oo_debt_to_income": {
            "qwnd_trailing_four_quarter": trailing_four_quarter_qwnd,
            "quarterly_values": oo_dti_quarterly_values,
            "annual_mean": float(fmean(oo_dti_quarterly_values)),
            "target_band": _band(min(oo_dti_quarterly_values), max(oo_dti_quarterly_values)),
        },
        "source_market_hpi": {
            "hpi_target_tolerance": HPI_TARGET_TOLERANCE,
            "index_sa_2024_values": hpi_2024_sa_values,
            "rebased_index_sa_2024_values": rebased_hpi_2024_values,
            "annual_mean": compute_rebased_mean(hpi_2024_sa_values),
            "full_history_std_window": {
                "start": "2005-01",
                "end": "2024-12",
                "count": len(hpi_full_sa_values),
            },
            "full_history_std": compute_rebased_std(hpi_full_sa_values),
            "cycle_period_months": estimate_dominant_cycle_period_months(hpi_full_index_values),
            "mean_target_band": _band(
                HPI_2024_REBASED_MEAN * (1.0 - HPI_TARGET_TOLERANCE),
                HPI_2024_REBASED_MEAN * (1.0 + HPI_TARGET_TOLERANCE),
            ),
            "std_target_band": _band(
                HPI_FULL_HISTORY_REBASED_STD * (1.0 - HPI_TARGET_TOLERANCE),
                HPI_FULL_HISTORY_REBASED_STD * (1.0 + HPI_TARGET_TOLERANCE),
            ),
            "cycle_target_band": _band(
                HPI_2024_CYCLE_PERIOD_MONTHS * (1.0 - HPI_TARGET_TOLERANCE),
                HPI_2024_CYCLE_PERIOD_MONTHS * (1.0 + HPI_TARGET_TOLERANCE),
            ),
        },
        "source_market_rpi": {
            "geography": "Great Britain",
            "rebased_monthly_values": rpi_rebased_values,
            "annual_mean": compute_rebased_mean(rpi_rebased_values),
            "target_band": _band(min(RPI_2024_GB_REBASED_MONTHLY_VALUES), max(RPI_2024_GB_REBASED_MONTHLY_VALUES)),
        },
        "source_household_tenure": {
            "household_owning_share": {
                "raw_source_value": tenure_values["household_owning_share"],
                "normalized_source_value": tenure_values["household_owning_share"],
                "target_band": _band(64.5, 65.5),
            },
            "household_renting_share": {
                "raw_source_value": tenure_values["household_renting_share"],
                "normalized_source_value": tenure_values["household_renting_share"],
                "target_band": _band(18.5, 19.5),
            },
        },
        "metric_definitions_core": {
            metric_id: _serialize_core_metric(metric_id)
            for metric_id, metric in TARGETS_BY_ID.items()
            if metric.kind in {"core_indicator", "output_series"}
        },
        "metric_definitions_household_share": {
            metric_id: _serialize_household_share_metric(metric_id)
            for metric_id, metric in TARGETS_BY_ID.items()
            if metric.kind == "household_share"
        },
        "metric_definitions_household_jsd": {
            metric_id: _serialize_household_metric(metric_id)
            for metric_id, metric in TARGETS_BY_ID.items()
            if metric.kind == "household_jsd"
        },
        "methodology_household_jsd_acceptance_band": {
            "metric_ids": [
                "income_distribution_jsd",
                "housing_wealth_distribution_jsd",
                "financial_wealth_distribution_jsd",
            ],
            "target_band": {"lower": 0.0, "upper": 0.12},
        },
    }


def _compare_jsonish(expected: object, actual: object, *, path: str, errors: list[str], float_tol: float = 1e-12) -> None:
    if isinstance(expected, float) or isinstance(actual, float):
        if abs(float(expected) - float(actual)) > float_tol:
            errors.append(f"{path}: expected {expected} but found {actual}")
        return
    if isinstance(expected, dict) and isinstance(actual, dict):
        expected_keys = set(expected.keys())
        actual_keys = set(actual.keys())
        if expected_keys != actual_keys:
            errors.append(f"{path}: key mismatch expected={sorted(expected_keys)} actual={sorted(actual_keys)}")
            return
        for key in sorted(expected_keys):
            _compare_jsonish(expected[key], actual[key], path=f"{path}.{key}", errors=errors, float_tol=float_tol)
        return
    if isinstance(expected, list) and isinstance(actual, list):
        if len(expected) != len(actual):
            errors.append(f"{path}: length mismatch expected={len(expected)} actual={len(actual)}")
            return
        for index, (expected_item, actual_item) in enumerate(zip(expected, actual, strict=True)):
            _compare_jsonish(expected_item, actual_item, path=f"{path}[{index}]", errors=errors, float_tol=float_tol)
        return
    if expected != actual:
        errors.append(f"{path}: expected {expected!r} but found {actual!r}")


def _load_review_entries(repo_root: Path) -> list[dict[str, object]]:
    payload = json.loads((repo_root / REVIEW_LEDGER_PATH).read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise RuntimeError("Review ledger must contain a JSON list of entries")
    return payload


def run_catalog_review(repo_root: Path | None = None) -> CatalogReviewResult:
    repo_root = repo_root or _repo_root()
    live_review_data = build_live_review_data(repo_root=repo_root)
    review_entries = _load_review_entries(repo_root)
    errors: list[str] = []

    live_review_ids = set(live_review_data.keys())
    ledger_review_ids = {str(entry["review_id"]) for entry in review_entries}
    if live_review_ids != ledger_review_ids:
        errors.append(
            f"Ledger coverage mismatch expected={sorted(live_review_ids)} actual={sorted(ledger_review_ids)}"
        )

    valid_statuses = {"verified", "needs_repo_snapshot", "needs_methodology_note"}
    all_metric_ids = set(TARGETS_BY_ID.keys())
    covered_metric_ids: set[str] = set()
    source_metric_ids: set[str] = {
        metric_id for metric_id, metric in TARGETS_BY_ID.items() if metric.source_metadata is not None
    }
    covered_source_metric_ids: set[str] = set()

    for entry in review_entries:
        review_id = str(entry["review_id"])
        status = str(entry["status"])
        if status not in valid_statuses:
            errors.append(f"{review_id}: invalid status {status}")
        for relative_path in entry.get("repo_source_paths", []):
            if not (repo_root / relative_path).exists():
                errors.append(f"{review_id}: missing repo source path {relative_path}")
        metric_ids = {str(metric_id) for metric_id in entry.get("metric_ids", [])}
        covered_metric_ids.update(metric_ids)
        if str(entry.get("review_category")) == "source_metadata":
            covered_source_metric_ids.update(metric_ids)
        expected_result = entry.get("expected_result")
        actual_result = live_review_data.get(review_id)
        if actual_result is None:
            continue
        _compare_jsonish(expected_result, actual_result, path=review_id, errors=errors)

    if covered_metric_ids != all_metric_ids:
        errors.append(
            f"Metric coverage mismatch expected={sorted(all_metric_ids)} actual={sorted(covered_metric_ids)}"
        )
    if covered_source_metric_ids != source_metric_ids:
        errors.append(
            f"Source coverage mismatch expected={sorted(source_metric_ids)} actual={sorted(covered_source_metric_ids)}"
        )

    return CatalogReviewResult(success=not errors, errors=errors, live_review_data=live_review_data)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run the 2024 validation catalog review.")
    parser.add_argument("--json", action="store_true", help="Print the live review payload as JSON on success")
    args = parser.parse_args()

    result = run_catalog_review()
    if not result.success:
        for error in result.errors:
            print(error)
        raise SystemExit(1)

    if args.json:
        print(json.dumps(result.live_review_data, indent=2, sort_keys=True))
    else:
        print("validation_catalog_2024 review passed")


if __name__ == "__main__":
    main()
